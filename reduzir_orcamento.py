import json
import os
import requests
import openpyxl
import time
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.FileHandler("reduzir_orcamento.log"),
        logging.StreamHandler()
    ]
)

logs_list = None

def log_message(msg):
    mensagem = f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(mensagem)
    with open("run_log.txt", "a", encoding="utf-8") as f:
        f.write(mensagem + "\n")
    if logs_list is not None:
        logs_list.append(mensagem)
    logging.info(msg)

# Leitura do arquivo de configuração
CONFIG_FILE = "config.json"
if not os.path.exists(CONFIG_FILE):
    default_config = {
        "ACCESS_TOKEN": "",
        "AD_ACCOUNTS": [],
        "LIMITE_LUCRO_BAIXO": 1000,
        "PERCENTUAL_REDUCAO": 0.10,
        "MINIMO_ORCAMENTO": 100,
        "WHATSAPP_GROUP": "#ZIP - ROAS IMPERIO",
        "DATE_PRESET": "today"
    }
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(default_config, f, indent=4, ensure_ascii=False)
with open(CONFIG_FILE, "r", encoding="utf-8") as f:
    config = json.load(f)

ACCESS_TOKEN = config.get("ACCESS_TOKEN", "")
AD_ACCOUNTS = config.get("AD_ACCOUNTS", [])
SPREADSHEET_PATH = ("campanhas_lucro_reducao.xlsx")
LIMITE_LUCRO_BAIXO = float(config.get("LIMITE_LUCRO_BAIXO", 10000))
PERCENTUAL_REDUCAO = float(config.get("PERCENTUAL_REDUCAO", 0.50))
MINIMO_ORCAMENTO = float(config.get("MINIMO_ORCAMENTO", 100))
WHATSAPP_GROUP = config.get("WHATSAPP_GROUP", "#ZIP - ROAS IMPERIO")
DATE_PRESET = config.get("DATE_PRESET", "today")

def criar_planilha():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "CAMPANHAS"
    sheet.append([
        "ID da Conta de Anúncio", "ID da Campanha", "Nome da Campanha",
        "Orçamento Diário", "Gasto", "Valor de Conversões",
        "ROAS", "Lucro", "Novo Orçamento"
    ])
    workbook.save(SPREADSHEET_PATH)
    log_message("Planilha criada com sucesso.")

def limpar_planilha():
    if os.path.exists(SPREADSHEET_PATH):
        try:
            workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
            if "CAMPANHAS" in workbook.sheetnames:
                sheet = workbook["CAMPANHAS"]
                sheet.delete_rows(2, sheet.max_row)
            else:
                workbook.create_sheet("CAMPANHAS")
            workbook.save(SPREADSHEET_PATH)
            log_message("Planilha limpa no início da execução.")
        except Exception as e:
            log_message(f"[ERRO] Falha ao limpar a planilha: {e}")
    else:
        log_message("Planilha não encontrada. Criando nova planilha.")
        criar_planilha()

def buscar_dados_facebook(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        log_message(f"[ERRO] Falha ao buscar dados do Facebook: {e}")
        return {}

def buscar_todos_dados_facebook(url):
    todos_dados = []
    while url:
        dados = buscar_dados_facebook(url)
        if "error" in dados:
            log_message(f"[ERRO] Graph API retornou: {dados.get('error')}")
            break
        page_data = dados.get("data", [])
        todos_dados.extend(page_data)
        url = dados.get("paging", {}).get("next")
    return todos_dados

def processar_dados_campanhas(campanhas, insights, ad_account):
    campanhas_filtradas = []
    for campanha in campanhas:
        if campanha.get("status", "").upper().strip() == "ACTIVE":
            insight = next((i for i in insights if i.get("campaign_id") == campanha.get("id")), None)
            if insight:
                gasto = float(insight.get("spend", 0))
                valor_conversao = sum(
                    float(a.get("value", 0))
                    for a in insight.get("action_values", [])
                    if a.get("action_type") in ['offsite_conversion.purchase', 'offsite_conversion.fb_pixel_purchase']
                )
            else:
                gasto = 0.0
                valor_conversao = 0.0
            daily_budget = float(campanha.get("daily_budget", 0)) / 100
            lucro = valor_conversao - gasto
            roas = round(valor_conversao / gasto, 2) if gasto > 0 else 0
            campanhas_filtradas.append({
                "id_conta": ad_account,
                "id_campanha": campanha["id"],
                "nome_campanha": campanha["name"],
                "orcamento_diario": daily_budget,
                "gasto": gasto,
                "valor_conversao": valor_conversao,
                "roas": roas,
                "lucro": lucro
            })
    return campanhas_filtradas

def salvar_campanhas_excel(campanhas):
    if not os.path.exists(SPREADSHEET_PATH):
        criar_planilha()
    try:
        workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
        if "CAMPANHAS" not in workbook.sheetnames:
            sheet = workbook.create_sheet("CAMPANHAS")
            sheet.append([
                "ID da Conta de Anúncio", "ID da Campanha", "Nome da Campanha",
                "Orçamento Diário", "Gasto", "Valor de Conversões",
                "ROAS", "Lucro", "Novo Orçamento"
            ])
        else:
            sheet = workbook["CAMPANHAS"]
        for campanha in campanhas:
            sheet.append([
                campanha["id_conta"],
                campanha["id_campanha"],
                campanha["nome_campanha"],
                campanha["orcamento_diario"],
                campanha["gasto"],
                campanha["valor_conversao"],
                campanha["roas"],
                campanha["lucro"],
                ""
            ])
        workbook.save(SPREADSHEET_PATH)
        log_message("Dados das campanhas salvos na planilha.")
    except Exception as e:
        log_message(f"[ERRO] Falha ao salvar planilha: {e}")

def calcular_orcamento_total():
    try:
        workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
        sheet = workbook["CAMPANHAS"]
        total = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            novo_orcamento = row[8] if row[8] is not None else row[3]
            total += novo_orcamento
        return total
    except Exception as e:
        log_message(f"[ERRO] Falha ao calcular orçamento total: {e}")
        return 0

def reduzir_campanhas():
    if not os.path.exists(SPREADSHEET_PATH):
        log_message("[ERRO] Planilha de campanhas não encontrada.")
        return
    workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
    if "CAMPANHAS" not in workbook.sheetnames:
        log_message("[ERRO] Aba 'CAMPANHAS' não encontrada na planilha.")
        return
    sheet = workbook["CAMPANHAS"]
    total_reduzido = 0.0
    # Itera pelas linhas a partir da segunda (ignorando cabeçalho)
    campanhas_reduzidas = 0
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        lucro = row[7]      # Coluna 8: Lucro
        orcamento = row[3]  # Coluna 4: Orçamento Diário
        if lucro is None or orcamento is None:
            continue
        if lucro < LIMITE_LUCRO_BAIXO:
            novo_orcamento = max(orcamento * (1 - PERCENTUAL_REDUCAO), MINIMO_ORCAMENTO)
            reducao = orcamento - novo_orcamento
            total_reduzido += reducao
            sheet.cell(row=row_index, column=9).value = novo_orcamento
            log_message(f"Campanha {row[1]} será reduzida de R$ {orcamento:.2f} para R$ {novo_orcamento:.2f}")
            resultado = atualizar_orcamento_facebook(row[1], novo_orcamento)
            if resultado:
                log_message(f"Campanha {row[1]} reduzida com sucesso para R$ {novo_orcamento:.2f}")
                campanhas_reduzidas += 1
    workbook.save(SPREADSHEET_PATH)
    total_orcamento_atual = calcular_orcamento_total()
    mensagem = (
        f"Redução realizada: {int(PERCENTUAL_REDUCAO * 100)}% dos orçamentos de campanhas com lucro abaixo de R$ {LIMITE_LUCRO_BAIXO} foram reduzidos. "
        f"Total de {campanhas_reduzidas} campanhas reduzidas. Valor total reduzido: R$ {total_reduzido:.2f}. Orçamento atual total: R$ {total_orcamento_atual:.2f}."
    )
    sucesso_whatsapp = enviar_mensagem_whatsapp(WHATSAPP_GROUP, mensagem)
    if not sucesso_whatsapp:
        log_message(f"[AVISO] Falha ao enviar mensagem WhatsApp, mas a redução foi concluída: {mensagem}")
    else:
        log_message(f"[INFO] Mensagem enviada com sucesso: {mensagem}")
    log_message("Processo de redução concluído com sucesso!")
    return True

def atualizar_orcamento_facebook(id_campanha, novo_orcamento):
    url = f"https://graph.facebook.com/v17.0/{id_campanha}"
    payload = {
        "daily_budget": int(novo_orcamento * 100),
        "access_token": ACCESS_TOKEN
    }
    try:
        log_message(f"Enviando atualização de orçamento para campanha {id_campanha}")
        response = requests.post(url, data=payload)
        result = response.json()
        log_message(f"Resposta da API: {result}")
        if result.get("success"):
            log_message(f"Orçamento atualizado para a campanha {id_campanha}: R$ {novo_orcamento:.2f}")
            return True
        else:
            log_message(f"[ERRO] Falha ao atualizar campanha {id_campanha}: {result.get('error', {}).get('message')}")
            return False
    except requests.exceptions.RequestException as e:
        log_message(f"[ERRO] Erro na requisição para atualizar orçamento: {e}")
        return False

def enviar_mensagem_whatsapp(grupo, mensagem):
    """
    Função atualizada para enviar mensagens para um grupo do WhatsApp Web/Business
    usando técnicas mais robustas de localização de elementos.
    
    Args:
        grupo (str): Nome do grupo para enviar a mensagem
        mensagem (str): Texto da mensagem a ser enviada
        
    Returns:
        bool: True se a mensagem foi enviada com sucesso, False caso contrário
    """
    log_message(f"Iniciando envio de mensagem para o grupo: {grupo}")
    driver = None
    
    try:
        # Configuração do navegador Chrome/Brave
        brave_options = Options()
        brave_options.binary_location = r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"
        brave_options.add_argument(r"--user-data-dir=C:\Users\Pichau\AppData\Local\BraveSoftware\Brave-Browser\User Data")
        brave_options.add_argument(r"--profile-directory=Default")
        brave_options.add_argument("--start-maximized")
        # Adicionar opção para evitar detecção de automação
        brave_options.add_argument("--disable-blink-features=AutomationControlled")
        brave_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        brave_options.add_experimental_option("useAutomationExtension", False)
        
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=brave_options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        # Abrir WhatsApp Web
        log_message("Abrindo WhatsApp Web...")
        driver.get("https://web.whatsapp.com")
        
        # Múltiplos seletores para verificar se o WhatsApp Web carregou
        whatsapp_loaded_selectors = [
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Caixa de texto de pesquisa']"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-placeholder='Pesquisar ou começar uma nova conversa']"),
            (By.XPATH, "//div[@role='textbox' and contains(@aria-label, 'Pesquisar')]"),
            (By.XPATH, "//div[contains(@class, '_ai04')]"),
            (By.XPATH, "//div[@id='pane-side']"),
            (By.ID, "pane-side")
        ]
        
        # Aguardar carregamento da página com múltiplos seletores
        log_message("Aguardando carregamento do WhatsApp Web...")
        carregou = False
        for selector_type, selector_value in whatsapp_loaded_selectors:
            try:
                WebDriverWait(driver, 60).until(
                    EC.presence_of_element_located((selector_type, selector_value))
                )
                log_message(f"WhatsApp Web carregado, elemento detectado: {selector_value}")
                carregou = True
                break
            except:
                continue
                
        # Verificação adicional se o WhatsApp carregou
        if not carregou:
            log_message("Verificando se o WhatsApp carregou usando outro método...")
            try:
                pane_side = WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@data-testid='chat-list']"))
                )
                log_message("WhatsApp Web carregado (detectado via chat-list)")
                carregou = True
            except:
                pass
                
        if not carregou:
            log_message("Falha ao detectar carregamento do WhatsApp Web")
            driver.save_screenshot("whatsapp_nao_carregou.png")
            return False
            
        time.sleep(2)  # Pequena pausa para garantir que a interface esteja pronta
        
        # Múltiplos seletores para o campo de pesquisa
        search_selectors = [
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Caixa de texto de pesquisa']"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-placeholder='Pesquisar ou começar uma nova conversa']"),
            (By.XPATH, "//div[@role='textbox' and contains(@aria-label, 'Pesquisar')]"),
            (By.XPATH, "//button[@aria-label='Pesquisar ou começar uma nova conversa']"),
            (By.XPATH, "//div[contains(@class, 'x10l6tqk')]"),
            (By.XPATH, "//div[contains(@class, 'lexical-rich-text-input')]//div[@role='textbox']"),
            (By.XPATH, "//div[@data-tab='3' and @role='textbox']")
        ]
        
        # Tentar encontrar e clicar no campo de pesquisa
        search_element = None
        for selector_type, selector_value in search_selectors:
            try:
                search_element = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((selector_type, selector_value))
                )
                log_message(f"Campo de pesquisa encontrado com seletor: {selector_value}")
                break
            except:
                continue
        
        if not search_element:
            # Tentar clicar na área onde geralmente fica o campo de pesquisa
            try:
                pane_side = driver.find_element(By.ID, "pane-side")
                driver.execute_script("arguments[0].scrollIntoView();", pane_side)
                actions = webdriver.ActionChains(driver)
                actions.move_to_element_with_offset(pane_side, 150, -50).click().perform()
                time.sleep(2)
                log_message("Clique realizado na área de pesquisa via coordenadas")
                
                # Tentar novamente encontrar o campo após o clique
                for selector_type, selector_value in search_selectors:
                    try:
                        search_element = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((selector_type, selector_value))
                        )
                        if search_element:
                            log_message(f"Campo de pesquisa encontrado após clique: {selector_value}")
                            break
                    except:
                        continue
            except Exception as e:
                log_message(f"Falha ao clicar na área de pesquisa: {e}")
                driver.save_screenshot("area_pesquisa_nao_encontrada.png")
        
        # Verificação adicional do campo de pesquisa usando outra abordagem
        if not search_element:
            try:
                # Tentar encontrar o campo de pesquisa por outros seletores mais genéricos
                search_elements = driver.find_elements(By.XPATH, "//div[@role='textbox']")
                if search_elements:
                    for elem in search_elements:
                        try:
                            placeholder = elem.get_attribute("aria-placeholder")
                            aria_label = elem.get_attribute("aria-label")
                            if (placeholder and "pesquisa" in placeholder.lower()) or \
                               (aria_label and "pesquisa" in aria_label.lower()):
                                search_element = elem
                                log_message("Campo de pesquisa encontrado via atributos aria")
                                break
                        except:
                            continue
            except:
                pass
        
        # Inserir texto de pesquisa
        if search_element:
            try:
                driver.execute_script("arguments[0].click();", search_element)
                time.sleep(1)
                search_element.clear()
                search_element.send_keys(grupo)
                log_message(f"Texto '{grupo}' inserido no campo de pesquisa")
            except Exception as e:
                log_message(f"Erro ao inserir texto no campo de pesquisa: {e}")
                # Tentar via ActionChains como alternativa
                actions = webdriver.ActionChains(driver)
                actions.move_to_element(search_element).click().send_keys(grupo).perform()
                log_message("Texto inserido via ActionChains")
        else:
            # Se ainda não encontrou o elemento, tentar enviar as teclas diretamente
            actions = webdriver.ActionChains(driver)
            actions.send_keys(grupo)
            actions.perform()
            log_message("Texto enviado diretamente via ActionChains")
        
        log_message(f"Pesquisando pelo grupo: {grupo}")
        time.sleep(5)  # Aguardar a pesquisa carregar resultados (tempo maior)
        
        # Múltiplos seletores para encontrar o grupo
        group_selectors = [
            (By.XPATH, f"//span[@title='{grupo}']"),
            (By.XPATH, f"//span[contains(text(), '{grupo}')]"),
            (By.XPATH, f"//div[contains(text(), '{grupo}')]"),
            (By.CSS_SELECTOR, f"div[aria-selected='true'] span[title='{grupo}']"),
            (By.CSS_SELECTOR, f"div[aria-selected='true'] div[title='{grupo}']"),
            (By.XPATH, f"//div[contains(@class, '_21S-L')]//span[@title='{grupo}']"),
            (By.XPATH, f"//div[@data-testid='cell-frame-title']//span[contains(text(), '{grupo}')]")
        ]
        
        # Tentar encontrar e clicar no grupo
        group_element = None
        for selector_type, selector_value in group_selectors:
            try:
                group_element = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((selector_type, selector_value))
                )
                log_message(f"Grupo encontrado com seletor: {selector_value}")
                break
            except:
                continue
        
        if group_element:
            # Tentar clicar via JavaScript (mais confiável)
            try:
                driver.execute_script("arguments[0].click();", group_element)
                log_message("Clicado no grupo usando JavaScript")
            except:
                try:
                    # Se falhar, tentar clique normal
                    group_element.click()
                    log_message("Clicado no grupo com método padrão")
                except Exception as e:
                    log_message(f"Erro ao clicar no grupo: {e}")
                    # Tentar clicar via ActionChains
                    actions = webdriver.ActionChains(driver)
                    actions.move_to_element(group_element).click().perform()
                    log_message("Clicado no grupo via ActionChains")
        else:
            # Tentar clicar no primeiro resultado da pesquisa
            try:
                first_result_selectors = [
                    (By.CSS_SELECTOR, "div[aria-selected='true']"),
                    (By.XPATH, "//div[@data-testid='cell-frame-container']"),
                    (By.XPATH, "//div[@role='row' and @aria-selected='true']"),
                    (By.XPATH, "//div[@data-testid='chat-list']//div[@role='button']")
                ]
                
                for selector_type, selector_value in first_result_selectors:
                    try:
                        first_result = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((selector_type, selector_value))
                        )
                        driver.execute_script("arguments[0].click();", first_result)
                        log_message(f"Clicado no primeiro resultado da pesquisa usando seletor: {selector_value}")
                        break
                    except:
                        continue
            except Exception as e:
                log_message(f"Grupo não encontrado: {e}")
                driver.save_screenshot("grupo_nao_encontrado.png")
                raise Exception(f"Grupo/contato '{grupo}' não encontrado")
        
        time.sleep(5)  # Pausa mais longa após clicar no grupo
        
        # Múltiplos seletores para o campo de mensagem
        message_selectors = [
            (By.XPATH, "//div[@role='textbox' and @data-tab='10']"),
            (By.XPATH, "//div[@role='textbox' and contains(@aria-label, 'Digite uma mensagem')]"),
            (By.CSS_SELECTOR, "div[role='textbox'][data-tab='10']"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Digite uma mensagem']"),
            (By.XPATH, "//div[contains(@class, 'lexical-rich-text-input')]//div[@role='textbox']"),
            (By.XPATH, "//footer//div[@role='textbox']"),
            (By.XPATH, "//div[contains(@class, '_3Uu1_')]"),
            (By.XPATH, "//div[@data-testid='conversation-compose-box-input']"),
            (By.XPATH, "//div[@title='Digite uma mensagem']")
        ]
        
        # Tentar encontrar o campo de mensagem
        log_message("Localizando campo de mensagem...")
        message_box = None
        for selector_type, selector_value in message_selectors:
            try:
                message_box = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((selector_type, selector_value))
                )
                log_message(f"Campo de mensagem encontrado com seletor: {selector_value}")
                break
            except:
                continue
        
        if not message_box:
            # Tentar encontrar qualquer footer ou área de mensagem na parte inferior
            try:
                # Primeiro, tentar encontrar o footer
                footers = driver.find_elements(By.TAG_NAME, "footer")
                if footers:
                    footer = footers[0]
                    driver.execute_script("arguments[0].scrollIntoView();", footer)
                    
                    # Tentar clicar em um ponto relativo ao footer
                    actions = webdriver.ActionChains(driver)
                    actions.move_to_element_with_offset(footer, 200, -30).click().perform()
                    log_message("Clique realizado na área de mensagem via coordenadas")
                    time.sleep(2)
                    
                    # Tentar novamente após o clique
                    for selector_type, selector_value in message_selectors:
                        try:
                            message_box = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((selector_type, selector_value))
                            )
                            if message_box:
                                log_message(f"Campo de mensagem encontrado após clique: {selector_value}")
                                break
                        except:
                            continue
                else:
                    # Procurar por elementos que possam ser o campo de composição
                    compose_elements = driver.find_elements(By.XPATH, "//div[@role='textbox']")
                    if compose_elements:
                        for elem in compose_elements:
                            try:
                                placeholder = elem.get_attribute("aria-placeholder")
                                aria_label = elem.get_attribute("aria-label")
                                if (placeholder and "mensagem" in placeholder.lower()) or \
                                   (aria_label and "mensagem" in aria_label.lower()):
                                    message_box = elem
                                    log_message("Campo de mensagem encontrado via atributos aria")
                                    break
                            except:
                                continue
                    
                    if not message_box:
                        # Tentar clicar em uma posição absoluta no centro-inferior da tela
                        window_size = driver.get_window_size()
                        x_position = window_size['width'] // 2
                        y_position = window_size['height'] - 100
                        actions = webdriver.ActionChains(driver)
                        actions.move_by_offset(x_position, y_position).click().perform()
                        log_message("Clique realizado em posição absoluta na tela")
                        time.sleep(2)
            except Exception as e:
                log_message(f"Falha ao localizar área de mensagem: {e}")
                driver.save_screenshot("area_mensagem_nao_encontrada.png")
        
        # Digitar a mensagem
        log_message("Tentando digitar mensagem...")
        if message_box:
            # Tentar clicar e limpar o campo antes de digitar
            try:
                driver.execute_script("arguments[0].click();", message_box)
                time.sleep(1)
                message_box.clear()
                message_box.send_keys(mensagem)
                log_message("Mensagem digitada com sucesso")
            except Exception as e:
                log_message(f"Erro ao digitar mensagem normalmente: {e}")
                # Tentar via ActionChains
                actions = webdriver.ActionChains(driver)
                actions.move_to_element(message_box).click().send_keys(mensagem).perform()
                log_message("Mensagem digitada via ActionChains")
        else:
            # Tentar enviar as teclas diretamente
            actions = webdriver.ActionChains(driver)
            actions.send_keys(mensagem)
            actions.perform()
            log_message("Mensagem digitada via ActionChains diretas")
        
        time.sleep(2)  # Pequena pausa após digitar a mensagem
        
        # Múltiplos seletores para o botão de enviar
        send_selectors = [
            (By.CSS_SELECTOR, "button[aria-label='Enviar']"),
            (By.CSS_SELECTOR, "button[data-tab='11']"),
            (By.XPATH, "//button[contains(@class, '_3wFFT')]"),
            (By.XPATH, "//button[@data-icon='send']"),
            (By.XPATH, "//button[contains(@class, '_1Ae7k')]"),
            (By.XPATH, "//span[@data-icon='send']"),
            (By.XPATH, "//button[@data-testid='send']"),
            (By.XPATH, "//button[@aria-label='Enviar mensagem']")
        ]
        
        # Tentar encontrar e clicar no botão de enviar
        log_message("Procurando botão de enviar...")
        send_button = None
        for selector_type, selector_value in send_selectors:
            try:
                send_button = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((selector_type, selector_value))
                )
                log_message(f"Botão de enviar encontrado com seletor: {selector_value}")
                break
            except:
                continue
        
        if send_button:
            # Tentar clicar via JavaScript (mais confiável)
            try:
                driver.execute_script("arguments[0].click();", send_button)
                log_message("Botão de enviar clicado via JavaScript")
            except Exception as e:
                log_message(f"Erro ao clicar via JavaScript: {e}")
                try:
                    # Se falhar, tentar clique normal
                    send_button.click()
                    log_message("Botão de enviar clicado com método padrão")
                except Exception as e2:
                    log_message(f"Erro ao clicar normalmente: {e2}")
                    # Tentar via ActionChains
                    actions = webdriver.ActionChains(driver)
                    actions.move_to_element(send_button).click().perform()
                    log_message("Botão clicado via ActionChains")
        else:
            # Tentar enviar com ENTER
            log_message("Tentando enviar mensagem com ENTER...")
            try:
                actions = webdriver.ActionChains(driver)
                actions.send_keys(Keys.ENTER)
                actions.perform()
                log_message("Tecla ENTER enviada")
            except Exception as e:
                log_message(f"Erro ao enviar tecla ENTER: {e}")
                if message_box:
                    try:
                        message_box.send_keys(Keys.ENTER)
                        log_message("ENTER enviado diretamente para o campo de mensagem")
                    except:
                        log_message("Não foi possível enviar ENTER para o campo de mensagem")
        
        # Aguardar um pouco para garantir que a mensagem foi enviada
        time.sleep(10)  # Tempo maior para garantir
        
        # Verificar se a mensagem foi enviada (procurar por marca de verificação)
        try:
            time.sleep(5)  # Tempo adicional para que a mensagem seja processada
            # Tentar localizar elementos que indicam mensagem enviada (check marks)
            check_marks = driver.find_elements(By.XPATH, "//span[@data-icon='msg-check']")
            double_check_marks = driver.find_elements(By.XPATH, "//span[@data-icon='msg-dcheck']")
            
            if check_marks or double_check_marks:
                log_message("Mensagem enviada com sucesso (confirmado por marcadores)")
            else:
                # Tentar encontrar outros indicadores de mensagem enviada
                sent_messages = driver.find_elements(By.XPATH, "//div[contains(@class, 'message-out')]")
                if sent_messages:
                    log_message("Mensagem possivelmente enviada (encontradas mensagens enviadas)")
                else:
                    log_message("Não foi possível confirmar se a mensagem foi enviada")
        except Exception as e:
            log_message(f"Erro ao verificar se a mensagem foi enviada: {e}")
        
        log_message("Processo de envio de mensagem concluído")
        # Considera que tentou enviar mesmo sem confirmação
        return True
        
    except Exception as e:
        log_message(f"Problema ao enviar mensagem no WhatsApp: {e}")
        if driver:
            try:
                driver.save_screenshot("whatsapp_error_final.png")
            except:
                pass
        return False
        
    finally:
        if driver:
            driver.quit()

def run(token, accounts, group, logs, date_range='today', start_date=None, end_date=None, low_profit=None, reduce_pct=None):
    global ACCESS_TOKEN, AD_ACCOUNTS, WHATSAPP_GROUP, DATE_PRESET, LIMITE_LUCRO_BAIXO, PERCENTUAL_REDUCAO, logs_list
    ACCESS_TOKEN = token
    AD_ACCOUNTS = accounts if accounts is not None else []
    WHATSAPP_GROUP = group
    logs_list = logs
    if date_range == 'custom' and start_date and end_date:
        DATE_PRESET = None
    else:
        presets = {'today': 'today', 'yesterday': 'yesterday', 'last7': 'last_7d'}
        DATE_PRESET = presets.get(date_range, 'today')
    if low_profit is not None:
        LIMITE_LUCRO_BAIXO = float(low_profit)
    if reduce_pct is not None:
        PERCENTUAL_REDUCAO = float(reduce_pct) if float(reduce_pct) <= 1.0 else float(reduce_pct)/100.0
    
    log_message("Iniciando execução da função run...")
    log_message("Parâmetros da operação:")
    log_message(f"- Token: {token[:5]}...{token[-5:]} (truncado)")
    log_message(f"- Contas: {accounts}")
    log_message(f"- Grupo: {group}")
    log_message(f"- Date Range: {date_range}")
    log_message(f"- Limite de Lucro Baixo: R$ {LIMITE_LUCRO_BAIXO}")
    log_message(f"- Percentual de Redução: {PERCENTUAL_REDUCAO * 100}%")
    
    limpar_planilha()
    todas_campanhas = []
    for ad_account in AD_ACCOUNTS:
        log_message(f"Processando conta de anúncio: {ad_account}")
        log_message(f"Buscando campanhas para conta {ad_account}...")
        campaigns_url = f"https://graph.facebook.com/v17.0/{ad_account}/campaigns?fields=id,name,daily_budget,status&access_token={ACCESS_TOKEN}"
        if date_range == 'custom' and start_date and end_date:
            insights_url = (
                f"https://graph.facebook.com/v17.0/{ad_account}/insights?fields=campaign_id,campaign_name,spend,action_values"
                f"&time_range[since]={start_date}&time_range[until]={end_date}&level=campaign&access_token={ACCESS_TOKEN}"
            )
        else:
            insights_url = (
                f"https://graph.facebook.com/v17.0/{ad_account}/insights?fields=campaign_id,campaign_name,spend,action_values"
                f"&date_preset={DATE_PRESET}&level=campaign&access_token={ACCESS_TOKEN}"
            )
        campaigns = buscar_todos_dados_facebook(campaigns_url)
        log_message(f"Encontradas {len(campaigns)} campanhas.")
        insights = buscar_todos_dados_facebook(insights_url)
        log_message(f"Encontrados {len(insights)} insights.")
        campanhas_processadas = processar_dados_campanhas(campaigns, insights, ad_account)
        log_message(f"Processadas {len(campanhas_processadas)} campanhas ativas.")
        todas_campanhas.extend(campanhas_processadas)
    
    log_message(f"Total de {len(todas_campanhas)} campanhas ativas encontradas.")
    salvar_campanhas_excel(todas_campanhas)
    log_message("Iniciando processo de redução...")
    
    try:
        reduzir_campanhas()
    except Exception as e:
        log_message(f"Erro geral ao reduzir campanhas: {e}")
        return False
    
    return True

if __name__ == "__main__":
    limpar_planilha()
    todas_campanhas = []
    for ad_account in AD_ACCOUNTS:
        log_message(f"[INFO] Processando conta de anúncio: {ad_account}")
        campaigns_url = f"https://graph.facebook.com/v17.0/{ad_account}/campaigns?fields=id,name,daily_budget,status&access_token={ACCESS_TOKEN}"
        insights_url = f"https://graph.facebook.com/v17.0/{ad_account}/insights?fields=campaign_id,campaign_name,spend,action_values&date_preset={DATE_PRESET}&level=campaign&access_token={ACCESS_TOKEN}"
        campaigns = buscar_todos_dados_facebook(campaigns_url)
        insights = buscar_todos_dados_facebook(insights_url)
        campanhas_processadas = processar_dados_campanhas(campaigns, insights, ad_account)
        todas_campanhas.extend(campanhas_processadas)
    salvar_campanhas_excel(todas_campanhas)
    reduzir_campanhas()