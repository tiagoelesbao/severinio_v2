import json
import os
import requests
import openpyxl
import time
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.FileHandler("realocar_orcamento.log"),
        logging.StreamHandler()
    ]
)

logs_list = None

def log_message(msg):
    mensagem = f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(mensagem)
    with open("run_log.txt", "a", encoding="utf-8") as f:
        f.write(mensagem + "\n")
    if logs_list is not None:
        logs_list.append(mensagem)
    logging.info(msg)

# Leitura de config.json
CONFIG_FILE = "config.json"
if not os.path.exists(CONFIG_FILE):
    default_config = {
        "ACCESS_TOKEN": "",
        "AD_ACCOUNTS": [],
        "ABO_ACCOUNTS": [],
        "LIMITE_LUCRO_BAIXO": 1000,
        "LIMITE_LUCRO_ALTO": 5000,
        "PERCENTUAL_REALOCACAO": 0.30,
        "MINIMO_ORCAMENTO": 100,
        "MAXIMO_ORCAMENTO": 10000,
        "WHATSAPP_GROUP": "#ZIP - ROAS IMPERIO",
        "DATE_PRESET": "today"
    }
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(default_config, f, indent=4, ensure_ascii=False)

with open(CONFIG_FILE, "r", encoding="utf-8") as f:
    config = json.load(f)

ACCESS_TOKEN = config.get("ACCESS_TOKEN", "")
AD_ACCOUNTS = config.get("AD_ACCOUNTS", [])
ABO_ACCOUNTS = config.get("ABO_ACCOUNTS", [])
SPREADSHEET_PATH = "campanhas_realocacao.xlsx"
LIMITE_LUCRO_BAIXO = float(config.get("LIMITE_LUCRO_BAIXO", 1000))
LIMITE_LUCRO_ALTO = float(config.get("LIMITE_LUCRO_ALTO", 5000))
PERCENTUAL_REALOCACAO = float(config.get("PERCENTUAL_REALOCACAO", 0.30))
MINIMO_ORCAMENTO = float(config.get("MINIMO_ORCAMENTO", 100))
MAXIMO_ORCAMENTO = float(config.get("MAXIMO_ORCAMENTO", 10000))
WHATSAPP_GROUP = config.get("WHATSAPP_GROUP", "#ZIP - ROAS IMPERIO")
DATE_PRESET = config.get("DATE_PRESET", "today")

# Armazena dados completos das campanhas para uso na realocação
campanhas_completas_data = {}

def criar_planilha():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "CAMPANHAS"
    sheet.append([
        "ID da Conta de Anúncio", "ID da Campanha", "Nome da Campanha", "Tipo",
        "Orçamento Diário", "Gasto", "Valor de Conversões",
        "ROAS", "Lucro", "Novo Orçamento", "Classificação", "Detalhes AdSets"
    ])
    workbook.save(SPREADSHEET_PATH)
    log_message("Planilha criada com sucesso.")

def limpar_planilha():
    if os.path.exists(SPREADSHEET_PATH):
        try:
            workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
            if "CAMPANHAS" in workbook.sheetnames:
                sheet = workbook["CAMPANHAS"]
                sheet.delete_rows(2, sheet.max_row)
            else:
                workbook.create_sheet("CAMPANHAS")
            workbook.save(SPREADSHEET_PATH)
            log_message("Planilha limpa no início da execução.")
        except Exception as e:
            log_message(f"[ERRO] Falha ao limpar a planilha: {e}")
    else:
        log_message("Planilha não encontrada. Criando nova planilha.")
        criar_planilha()

def buscar_dados_facebook(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        log_message(f"[ERRO] Falha ao buscar dados do Facebook: {e}\nResposta: {response.text if 'response' in locals() else ''}")
        return {}

def buscar_todos_dados_facebook(url):
    todos_dados = []
    while url:
        dados = buscar_dados_facebook(url)
        if "error" in dados:
            log_message(f"[ERRO] Graph API retornou: {dados['error']}")
            break
        page_data = dados.get("data", [])
        todos_dados.extend(page_data)
        url = dados.get("paging", {}).get("next")
    return todos_dados

def detectar_tipo_campanha(campanha, ad_account):
    """Detecta se a campanha é CBO ou ABO"""
    daily_budget = campanha.get("daily_budget", 0)
    
    if ad_account in ABO_ACCOUNTS:
        return "ABO"
    
    if daily_budget and int(daily_budget) > 0:
        return "CBO"
    else:
        return "ABO"

def buscar_adsets_campanha(campaign_id):
    """Busca todos os ad sets de uma campanha ABO"""
    url = (
        f"https://graph.facebook.com/v17.0/{campaign_id}/adsets"
        f"?fields=id,name,daily_budget,status&access_token={ACCESS_TOKEN}"
    )
    return buscar_todos_dados_facebook(url)

def buscar_insights_adset(ad_account, campaign_id, date_preset=None, start_date=None, end_date=None):
    """Busca insights no nível de ad set para campanhas ABO"""
    filtering = f'[{{"field":"campaign_id","operator":"EQUAL","value":"{campaign_id}"}}]'
    
    if date_preset:
        url = (
            f"https://graph.facebook.com/v17.0/{ad_account}/insights"
            f"?fields=adset_id,adset_name,campaign_id,spend,action_values"
            f"&date_preset={date_preset}&level=adset"
            f"&filtering={filtering}"
            f"&access_token={ACCESS_TOKEN}"
        )
    else:
        url = (
            f"https://graph.facebook.com/v17.0/{ad_account}/insights"
            f"?fields=adset_id,adset_name,campaign_id,spend,action_values"
            f"&time_range[since]={start_date}&time_range[until]={end_date}"
            f"&level=adset"
            f"&filtering={filtering}"
            f"&access_token={ACCESS_TOKEN}"
        )
    
    return buscar_todos_dados_facebook(url)

def processar_campanha_abo(campanha, ad_account, date_preset=None, start_date=None, end_date=None):
    """Processa campanhas ABO agregando dados de todos os ad sets ativos"""
    campaign_id = campanha["id"]
    log_message(f"Processando campanha ABO: {campanha['name']}")
    
    adsets = buscar_adsets_campanha(campaign_id)
    insights_adsets = buscar_insights_adset(ad_account, campaign_id, date_preset, start_date, end_date)
    
    # Agregar dados de todos os ad sets ativos
    total_orcamento = 0
    total_gasto = 0
    total_conversao = 0
    adsets_info = []
    adsets_ativos = 0
    
    for adset in adsets:
        if adset.get("status") == "ACTIVE":
            adsets_ativos += 1
            adset_id = adset["id"]
            daily_budget = float(adset.get("daily_budget", 0)) / 100
            total_orcamento += daily_budget
            
            # Buscar insight correspondente
            insight = next((i for i in insights_adsets if i.get("adset_id") == adset_id), None)
            
            if insight:
                gasto = float(insight.get("spend", 0))
                valor_conversao = sum(
                    float(a.get("value", 0))
                    for a in insight.get("action_values", [])
                    if a.get("action_type") in ['offsite_conversion.purchase', 'offsite_conversion.fb_pixel_purchase']
                )
                total_gasto += gasto
                total_conversao += valor_conversao
            else:
                gasto = 0
                valor_conversao = 0
            
            lucro = valor_conversao - gasto
            
            adsets_info.append({
                "adset_id": adset_id,
                "adset_name": adset["name"],
                "daily_budget": daily_budget,
                "gasto": gasto,
                "valor_conversao": valor_conversao,
                "lucro": lucro
            })
    
    log_message(f"Campanha ABO {campaign_id}: {adsets_ativos} adsets ativos, orçamento total: R$ {total_orcamento:.2f}")
    
    lucro = total_conversao - total_gasto
    roas = round(total_conversao / total_gasto, 2) if total_gasto > 0 else 0
    
    # Classificação baseada no lucro
    if lucro < LIMITE_LUCRO_BAIXO:
        classificacao = "BAIXO"
    elif lucro >= LIMITE_LUCRO_ALTO:
        classificacao = "ALTO"
    else:
        classificacao = "MÉDIO"
    
    return {
        "id_conta": ad_account,
        "id_campanha": campaign_id,
        "nome_campanha": campanha["name"],
        "tipo_campanha": "ABO",
        "orcamento_diario": total_orcamento,
        "gasto": total_gasto,
        "valor_conversao": total_conversao,
        "roas": roas,
        "lucro": lucro,
        "classificacao": classificacao,
        "adsets_info": adsets_info,
        "detalhes_adsets": f"{adsets_ativos} adsets ativos"
    }

def processar_dados_campanhas(campanhas, insights, ad_account, date_preset=None, start_date=None, end_date=None):
    """Processa dados das campanhas, detectando automaticamente se são CBO ou ABO"""
    campanhas_filtradas = []
    
    for campanha in campanhas:
        if campanha.get("status", "").upper().strip() == "ACTIVE":
            tipo_campanha = detectar_tipo_campanha(campanha, ad_account)
            
            if tipo_campanha == "ABO":
                # Processar como ABO
                dados_campanha = processar_campanha_abo(campanha, ad_account, date_preset, start_date, end_date)
                campanhas_filtradas.append(dados_campanha)
                # Armazenar dados completos para uso posterior
                campanhas_completas_data[campanha["id"]] = dados_campanha
            else:
                # Processar como CBO (código original)
                insight = next((i for i in insights if i.get("campaign_id") == campanha.get("id")), None)
                
                if insight:
                    gasto = float(insight.get("spend", 0))
                    valor_conversao = sum(
                        float(a.get("value", 0))
                        for a in insight.get("action_values", [])
                        if a.get("action_type") in ['offsite_conversion.purchase', 'offsite_conversion.fb_pixel_purchase']
                    )
                else:
                    gasto = 0.0
                    valor_conversao = 0.0
                
                daily_budget = float(campanha.get("daily_budget", 0)) / 100
                lucro = valor_conversao - gasto
                roas = round(valor_conversao / gasto, 2) if gasto > 0 else 0
                
                # Classificação baseada no lucro
                if lucro < LIMITE_LUCRO_BAIXO:
                    classificacao = "BAIXO"
                elif lucro >= LIMITE_LUCRO_ALTO:
                    classificacao = "ALTO"
                else:
                    classificacao = "MÉDIO"
                
                dados_campanha = {
                    "id_conta": ad_account,
                    "id_campanha": campanha["id"],
                    "nome_campanha": campanha["name"],
                    "tipo_campanha": "CBO",
                    "orcamento_diario": daily_budget,
                    "gasto": gasto,
                    "valor_conversao": valor_conversao,
                    "roas": roas,
                    "lucro": lucro,
                    "classificacao": classificacao,
                    "adsets_info": None,
                    "detalhes_adsets": "N/A"
                }
                
                campanhas_filtradas.append(dados_campanha)
                campanhas_completas_data[campanha["id"]] = dados_campanha
    
    return campanhas_filtradas

def salvar_campanhas_excel(campanhas):
    if not os.path.exists(SPREADSHEET_PATH):
        criar_planilha()
    
    try:
        workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
        sheet = workbook["CAMPANHAS"]
        
        for campanha in campanhas:
            sheet.append([
                campanha["id_conta"],
                campanha["id_campanha"],
                campanha["nome_campanha"],
                campanha.get("tipo_campanha", "CBO"),
                campanha["orcamento_diario"],
                campanha["gasto"],
                campanha["valor_conversao"],
                campanha["roas"],
                campanha["lucro"],
                "",  # Novo orçamento
                campanha["classificacao"],
                campanha.get("detalhes_adsets", "")
            ])
        
        workbook.save(SPREADSHEET_PATH)
        log_message(f"Dados de {len(campanhas)} campanhas salvos na planilha.")
    except Exception as e:
        log_message(f"[ERRO] Falha ao salvar planilha: {e}")

def atualizar_orcamento_adset(adset_id, novo_orcamento):
    """Atualiza o orçamento de um ad set específico"""
    url = f"https://graph.facebook.com/v17.0/{adset_id}"
    payload = {
        "daily_budget": int(novo_orcamento * 100),
        "access_token": ACCESS_TOKEN
    }
    
    try:
        log_message(f"Atualizando orçamento do AdSet {adset_id} para R$ {novo_orcamento:.2f}")
        response = requests.post(url, data=payload)
        result = response.json()
        
        if result.get("success"):
            log_message(f"Orçamento do AdSet {adset_id} atualizado com sucesso")
            return True
        else:
            erro_msg = result.get('error', {}).get('message', 'Erro desconhecido')
            log_message(f"[ERRO] Falha ao atualizar AdSet {adset_id}: {erro_msg}")
            return False
    except requests.exceptions.RequestException as e:
        log_message(f"[ERRO] Erro na requisição para atualizar AdSet {adset_id}: {e}")
        return False

def realocar_campanha_abo(campanha_info, operacao, percentual):
    """
    Realoca orçamento de uma campanha ABO
    operacao: 'reduzir' ou 'aumentar'
    percentual: percentual de mudança
    """
    adsets_info = campanha_info.get("adsets_info", [])
    if not adsets_info:
        log_message(f"[AVISO] Campanha ABO {campanha_info['id_campanha']} sem adsets ativos")
        return 0, 0
    
    total_mudanca = 0
    adsets_modificados = 0
    
    if operacao == 'reduzir':
        # Para redução, reduzir proporcionalmente todos os adsets
        for adset in adsets_info:
            reducao = adset['daily_budget'] * percentual
            novo_orcamento = max(adset['daily_budget'] - reducao, MINIMO_ORCAMENTO)
            reducao_real = adset['daily_budget'] - novo_orcamento
            
            if atualizar_orcamento_adset(adset['adset_id'], novo_orcamento):
                total_mudanca += reducao_real
                adsets_modificados += 1
                log_message(f"AdSet {adset['adset_name']} reduzido de R$ {adset['daily_budget']:.2f} para R$ {novo_orcamento:.2f}")
    
    else:  # aumentar
        # Para aumento, distribuir proporcionalmente ao lucro
        adsets_lucrativos = [a for a in adsets_info if a['lucro'] > 0]
        
        if adsets_lucrativos:
            total_lucro_adsets = sum(a['lucro'] for a in adsets_lucrativos)
            
            for adset in adsets_lucrativos:
                proporcao = adset['lucro'] / total_lucro_adsets
                incremento = percentual * campanha_info['orcamento_diario'] * proporcao
                novo_orcamento = min(adset['daily_budget'] + incremento, MAXIMO_ORCAMENTO)
                incremento_real = novo_orcamento - adset['daily_budget']
                
                if atualizar_orcamento_adset(adset['adset_id'], novo_orcamento):
                    total_mudanca += incremento_real
                    adsets_modificados += 1
                    log_message(f"AdSet {adset['adset_name']} aumentado de R$ {adset['daily_budget']:.2f} para R$ {novo_orcamento:.2f}")
        else:
            # Se nenhum adset for lucrativo, distribuir igualmente
            incremento_por_adset = (percentual * campanha_info['orcamento_diario']) / len(adsets_info)
            
            for adset in adsets_info:
                novo_orcamento = min(adset['daily_budget'] + incremento_por_adset, MAXIMO_ORCAMENTO)
                incremento_real = novo_orcamento - adset['daily_budget']
                
                if atualizar_orcamento_adset(adset['adset_id'], novo_orcamento):
                    total_mudanca += incremento_real
                    adsets_modificados += 1
    
    return total_mudanca, adsets_modificados

def atualizar_orcamento_facebook(id_campanha, novo_orcamento):
    """Atualiza o orçamento de campanhas CBO"""
    url = f"https://graph.facebook.com/v17.0/{id_campanha}"
    payload = {
        "daily_budget": int(novo_orcamento * 100),
        "access_token": ACCESS_TOKEN
    }
    try:
        log_message(f"Enviando atualização de orçamento para campanha {id_campanha}")
        response = requests.post(url, data=payload)
        result = response.json()
        log_message(f"Resposta da API: {result}")
        if result.get("success"):
            log_message(f"Orçamento atualizado para a campanha {id_campanha}: R$ {novo_orcamento:.2f}")
            return True
        else:
            log_message(f"[ERRO] Falha ao atualizar campanha {id_campanha}: {result.get('error', {}).get('message')}")
            return False
    except requests.exceptions.RequestException as e:
        log_message(f"[ERRO] Erro na requisição para atualizar orçamento: {e}")
        return False

def calcular_orcamento_total():
    try:
        workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
        sheet = workbook["CAMPANHAS"]
        total = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            novo_orcamento = row[9] if row[9] is not None and row[9] != "" else row[4]
            total += novo_orcamento
        return total
    except Exception as e:
        log_message(f"[ERRO] Falha ao calcular orçamento total: {e}")
        return 0

def realocar_orcamentos():
    """Realoca orçamentos entre campanhas baixas e altas, suportando CBO e ABO"""
    if not os.path.exists(SPREADSHEET_PATH):
        log_message("[ERRO] Planilha de campanhas não encontrada.")
        return False
    
    workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
    if "CAMPANHAS" not in workbook.sheetnames:
        log_message("[ERRO] Aba 'CAMPANHAS' não encontrada na planilha.")
        return False
    
    sheet = workbook["CAMPANHAS"]
    
    # Identificar campanhas com lucro baixo e alto
    campanhas_baixo = []
    campanhas_alto = []
    
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        classificacao = row[10]  # Classificação
        
        if classificacao == "BAIXO":
            campanhas_baixo.append({
                "linha_index": row_index,
                "id_campanha": row[1],
                "nome_campanha": row[2],
                "tipo_campanha": row[3] if len(row) > 3 else "CBO",
                "orcamento_diario": row[4],
                "lucro": row[8]
            })
        elif classificacao == "ALTO":
            campanhas_alto.append({
                "linha_index": row_index,
                "id_campanha": row[1],
                "nome_campanha": row[2],
                "tipo_campanha": row[3] if len(row) > 3 else "CBO",
                "orcamento_diario": row[4],
                "lucro": row[8]
            })
    
    log_message(f"Encontradas {len(campanhas_baixo)} campanhas com lucro BAIXO e {len(campanhas_alto)} com lucro ALTO")
    
    # Se não houver campanhas para reduzir ou aumentar, finalizar
    if not campanhas_baixo or not campanhas_alto:
        log_message("Não há campanhas suficientes para realocação.")
        return False
    
    # Calcular total a ser realocado
    total_reducao = 0
    campanhas_reduzidas = 0
    campanhas_cbo_reduzidas = 0
    campanhas_abo_reduzidas = 0
    total_adsets_reduzidos = 0
    
    # Reduzir orçamentos das campanhas com lucro baixo
    for campanha in campanhas_baixo:
        if campanha["tipo_campanha"] == "ABO":
            # Para ABO, realocar nos adsets
            campanha_completa = campanhas_completas_data.get(campanha["id_campanha"])
            if campanha_completa and campanha_completa.get("adsets_info"):
                mudanca, adsets_mod = realocar_campanha_abo(campanha_completa, 'reduzir', PERCENTUAL_REALOCACAO)
                if mudanca > 0:
                    total_reducao += mudanca
                    novo_orcamento_total = campanha["orcamento_diario"] - mudanca
                    sheet.cell(row=campanha["linha_index"], column=10).value = novo_orcamento_total
                    campanhas_reduzidas += 1
                    campanhas_abo_reduzidas += 1
                    total_adsets_reduzidos += adsets_mod
                    log_message(f"Campanha ABO {campanha['nome_campanha']} reduzida em R$ {mudanca:.2f}")
        else:
            # Para CBO, reduzir normalmente
            reducao = campanha["orcamento_diario"] * PERCENTUAL_REALOCACAO
            novo_orcamento = max(campanha["orcamento_diario"] - reducao, MINIMO_ORCAMENTO)
            reducao_real = campanha["orcamento_diario"] - novo_orcamento
            total_reducao += reducao_real
            
            sheet.cell(row=campanha["linha_index"], column=10).value = novo_orcamento
            
            if atualizar_orcamento_facebook(campanha["id_campanha"], novo_orcamento):
                log_message(f"Campanha CBO {campanha['nome_campanha']} reduzida de R$ {campanha['orcamento_diario']:.2f} para R$ {novo_orcamento:.2f}")
                campanhas_reduzidas += 1
                campanhas_cbo_reduzidas += 1
    
    # Distribuir o valor reduzido entre as campanhas com lucro alto
    if total_reducao > 0 and campanhas_alto:
        soma_lucro_alto = sum(c["lucro"] for c in campanhas_alto)
        campanhas_escaladas = 0
        campanhas_cbo_escaladas = 0
        campanhas_abo_escaladas = 0
        total_adsets_escalados = 0
        
        for campanha in campanhas_alto:
            # Distribuir proporcionalmente ao lucro
            proporcao = campanha["lucro"] / soma_lucro_alto if soma_lucro_alto > 0 else 1.0 / len(campanhas_alto)
            incremento_total = total_reducao * proporcao
            
            if campanha["tipo_campanha"] == "ABO":
                # Para ABO, distribuir entre adsets
                campanha_completa = campanhas_completas_data.get(campanha["id_campanha"])
                if campanha_completa and campanha_completa.get("adsets_info"):
                    mudanca, adsets_mod = realocar_campanha_abo(campanha_completa, 'aumentar', incremento_total / campanha["orcamento_diario"])
                    if mudanca > 0:
                        novo_orcamento_total = campanha["orcamento_diario"] + mudanca
                        sheet.cell(row=campanha["linha_index"], column=10).value = novo_orcamento_total
                        campanhas_escaladas += 1
                        campanhas_abo_escaladas += 1
                        total_adsets_escalados += adsets_mod
                        log_message(f"Campanha ABO {campanha['nome_campanha']} aumentada em R$ {mudanca:.2f}")
            else:
                # Para CBO, aumentar normalmente
                novo_orcamento = min(campanha["orcamento_diario"] + incremento_total, MAXIMO_ORCAMENTO)
                
                sheet.cell(row=campanha["linha_index"], column=10).value = novo_orcamento
                
                if atualizar_orcamento_facebook(campanha["id_campanha"], novo_orcamento):
                    log_message(f"Campanha CBO {campanha['nome_campanha']} aumentada de R$ {campanha['orcamento_diario']:.2f} para R$ {novo_orcamento:.2f}")
                    campanhas_escaladas += 1
                    campanhas_cbo_escaladas += 1
    
    # Salvar planilha
    workbook.save(SPREADSHEET_PATH)
    
    # Calcular orçamento total atual
    total_orcamento_atual = calcular_orcamento_total()
    
    # Preparar mensagem de relatório detalhada
    mensagem = (
        f"✅ Realocação concluída!\n\n"
        f"📉 Redução:\n"
        f"• {campanhas_reduzidas} campanhas ({campanhas_cbo_reduzidas} CBO, {campanhas_abo_reduzidas} ABO)\n"
    )
    
    if campanhas_abo_reduzidas > 0:
        mensagem += f"• {total_adsets_reduzidos} ad sets reduzidos\n"
    
    mensagem += (
        f"• Total reduzido: R$ {total_reducao:.2f}\n\n"
        f"📈 Aumento:\n"
        f"• {len(campanhas_alto)} campanhas ({campanhas_cbo_escaladas} CBO, {campanhas_abo_escaladas} ABO)\n"
    )
    
    if campanhas_abo_escaladas > 0:
        mensagem += f"• {total_adsets_escalados} ad sets aumentados\n"
    
    mensagem += (
        f"• Total distribuído: R$ {total_reducao:.2f}\n\n"
        f"💰 Orçamento total: R$ {total_orcamento_atual:.2f}"
    )
    
    # Enviar relatório via WhatsApp
    sucesso_whatsapp = enviar_mensagem_whatsapp(WHATSAPP_GROUP, mensagem)
    
    if not sucesso_whatsapp:
        log_message(f"[AVISO] Falha ao enviar mensagem WhatsApp, mas a realocação foi concluída")
    else:
        log_message(f"[INFO] Mensagem enviada com sucesso")
    
    log_message("Processo de realocação concluído com sucesso!")
    return True

def enviar_mensagem_whatsapp(grupo, mensagem):
    """Função para enviar mensagens para um grupo do WhatsApp"""
    log_message(f"Iniciando envio de mensagem para o grupo: {grupo}")
    driver = None
    
    try:
        # Configuração do navegador Chrome/Brave
        brave_options = Options()
        brave_options.binary_location = r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"
        brave_options.add_argument(r"--user-data-dir=C:\Users\Pichau\AppData\Local\BraveSoftware\Brave-Browser\User Data")
        brave_options.add_argument(r"--profile-directory=Default")
        brave_options.add_argument("--start-maximized")
        brave_options.add_argument("--window-size=1920,1080")
        
        # Adicionar opção para evitar detecção de automação
        brave_options.add_argument("--disable-blink-features=AutomationControlled")
        brave_options.add_argument("--disable-extensions")
        brave_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        brave_options.add_experimental_option("useAutomationExtension", False)
        
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=brave_options)
        
        # Remover flag navigator.webdriver
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            })
            """
        })
        
        # Abrir WhatsApp Web
        log_message("Abrindo WhatsApp Web...")
        driver.get("https://web.whatsapp.com")
        
        # Múltiplos seletores para verificar se o WhatsApp Web carregou
        whatsapp_loaded_selectors = [
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Caixa de texto de pesquisa']"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-placeholder='Pesquisar ou começar uma nova conversa']"),
            (By.XPATH, "//div[@role='textbox' and contains(@aria-label, 'Pesquisar')]"),
            (By.XPATH, "//div[contains(@class, '_ai04')]"),
            (By.XPATH, "//div[@id='pane-side']")
        ]
        
        # Aguardar carregamento da página com múltiplos seletores
        log_message("Aguardando carregamento do WhatsApp Web...")
        carregou = False
        for selector_type, selector_value in whatsapp_loaded_selectors:
            try:
                WebDriverWait(driver, 60).until(
                    EC.presence_of_element_located((selector_type, selector_value))
                )
                log_message(f"WhatsApp Web carregado, elemento detectado: {selector_value}")
                carregou = True
                break
            except:
                continue
        
        if not carregou:
            log_message("Verificando se o WhatsApp carregou usando outro método...")
            try:
                WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@data-testid='chat-list']"))
                )
                log_message("WhatsApp Web carregado (detectado via chat-list)")
                carregou = True
            except:
                pass
        
        if not carregou:
            log_message("Falha ao detectar carregamento do WhatsApp Web")
            driver.save_screenshot("whatsapp_nao_carregou.png")
            return False
        
        # Garantir que a página está completamente carregada
        time.sleep(2)
        
        # Múltiplos seletores para o campo de pesquisa
        search_selectors = [
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Caixa de texto de pesquisa']"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-placeholder='Pesquisar ou começar uma nova conversa']"),
            (By.XPATH, "//div[@role='textbox' and contains(@aria-label, 'Pesquisar')]"),
            (By.XPATH, "//button[@aria-label='Pesquisar ou começar uma nova conversa']"),
            (By.XPATH, "//div[contains(@class, 'x10l6tqk')]"),
            (By.XPATH, "//div[contains(@class, 'lexical-rich-text-input')]//div[@role='textbox']")
        ]
        
        # Encontrar o campo de pesquisa
        search_element = None
        for selector_type, selector_value in search_selectors:
            try:
                search_element = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((selector_type, selector_value))
                )
                log_message(f"Campo de pesquisa encontrado com seletor: {selector_value}")
                break
            except:
                continue
        
        if not search_element:
            log_message("Campo de pesquisa não encontrado")
            driver.save_screenshot("campo_pesquisa_nao_encontrado.png")
            return False
        
        # Clicar no campo de pesquisa e inserir o nome do grupo
        try:
            # Usando clique normal primeiro
            search_element.click()
            time.sleep(1)
            search_element.clear()
            search_element.send_keys(grupo)
            log_message(f"Texto de pesquisa '{grupo}' inserido.")
        except Exception as e:
            log_message(f"Erro ao clicar no campo de pesquisa normalmente: {e}")
            try:
                # Se falhar, tente com ActionChains
                actions = webdriver.ActionChains(driver)
                actions.move_to_element(search_element).click().perform()
                time.sleep(1)
                search_element.clear()
                search_element.send_keys(grupo)
                log_message(f"Texto de pesquisa '{grupo}' inserido via ActionChains.")
            except Exception as e2:
                log_message(f"Falha ao inserir texto de pesquisa: {e2}")
                driver.save_screenshot("erro_inserir_pesquisa.png")
                return False
        
        # Aguardar resultados da pesquisa
        time.sleep(3)
        
        # Múltiplos seletores para encontrar o grupo
        group_selectors = [
            (By.XPATH, f"//span[@title='{grupo}']"),
            (By.XPATH, f"//span[contains(text(), '{grupo}')]"),
            (By.XPATH, f"//div[contains(text(), '{grupo}')]")
        ]
        
        # Encontrar e clicar no grupo
        group_element = None
        for selector_type, selector_value in group_selectors:
            try:
                group_element = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((selector_type, selector_value))
                )
                log_message(f"Grupo encontrado com seletor: {selector_value}")
                break
            except:
                continue
        
        if not group_element:
            log_message("Grupo não encontrado")
            driver.save_screenshot("grupo_nao_encontrado.png")
            return False
        
        # Clicar no grupo usando clique normal
        try:
            group_element.click()
            log_message("Grupo clicado com sucesso")
        except Exception as e:
            log_message(f"Erro ao clicar no grupo normalmente: {e}")
            try:
                # Se falhar, tente com ActionChains
                actions = webdriver.ActionChains(driver)
                actions.move_to_element(group_element).click().perform()
                log_message("Grupo clicado com ActionChains")
            except Exception as e2:
                log_message(f"Falha ao clicar no grupo: {e2}")
                driver.save_screenshot("erro_clique_grupo.png")
                return False
        
        # Verificar se o grupo foi realmente clicado
        time.sleep(3)
        try:
            # Verificar se estamos na conversa, procurando por elemento que só aparece depois de clicar no grupo
            chat_header = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, f"//span[contains(text(), '{grupo}')]/ancestor::div[contains(@role, 'button')]"))
            )
            log_message("Conversa do grupo aberta com sucesso")
        except:
            log_message("Não foi possível confirmar se a conversa do grupo foi aberta")
            driver.save_screenshot("conversa_nao_confirmada.png")
            return False
        
        # Múltiplos seletores para o campo de mensagem
        message_selectors = [
            (By.XPATH, "//div[@role='textbox' and @data-tab='10']"),
            (By.XPATH, "//div[@role='textbox' and contains(@aria-label, 'Digite uma mensagem')]"),
            (By.CSS_SELECTOR, "div[role='textbox'][data-tab='10']"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Digite uma mensagem']"),
            (By.XPATH, "//div[contains(@class, 'lexical-rich-text-input')]//div[@role='textbox']"),
            (By.XPATH, "//footer//div[@role='textbox']"),
            (By.XPATH, "//div[@data-testid='conversation-compose-box-input']"),
            (By.XPATH, "//div[@title='Digite uma mensagem']")
        ]
        
        # Encontrar o campo de mensagem
        log_message("Localizando campo de mensagem...")
        message_box = None
        for selector_type, selector_value in message_selectors:
            try:
                message_box = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((selector_type, selector_value))
                )
                log_message(f"Campo de mensagem encontrado com seletor: {selector_value}")
                break
            except:
                continue
        
        if not message_box:
            log_message("Campo de mensagem não encontrado")
            driver.save_screenshot("campo_mensagem_nao_encontrado.png")
            return False
        
        # Clicar e digitar a mensagem
        log_message("Tentando digitar mensagem...")
        try:
            message_box.click()
            time.sleep(1)
            message_box.clear()
            message_box.send_keys(mensagem)
            log_message("Mensagem digitada com sucesso")
        except Exception as e:
            log_message(f"Erro ao digitar mensagem normalmente: {e}")
            try:
                actions = webdriver.ActionChains(driver)
                actions.move_to_element(message_box).click().perform()
                time.sleep(1)
                message_box.send_keys(mensagem)
                log_message("Mensagem digitada via ActionChains")
            except Exception as e2:
                log_message(f"Falha ao digitar mensagem: {e2}")
                driver.save_screenshot("erro_digitacao.png")
                return False
        
        time.sleep(1)
        
        # Enviar a mensagem
        log_message("Enviando mensagem...")
        
        # Primeiro tenta encontrar o botão de enviar
        send_selectors = [
            (By.CSS_SELECTOR, "button[aria-label='Enviar']"),
            (By.CSS_SELECTOR, "button[data-tab='11']"),
            (By.XPATH, "//button[contains(@class, '_3wFFT')]"),
            (By.XPATH, "//button[@data-icon='send']"),
            (By.XPATH, "//span[@data-icon='send']"),
            (By.XPATH, "//button[@data-testid='send']"),
            (By.XPATH, "//button[@aria-label='Enviar mensagem']")
        ]
        
        send_button = None
        for selector_type, selector_value in send_selectors:
            try:
                send_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((selector_type, selector_value))
                )
                log_message(f"Botão de enviar encontrado com seletor: {selector_value}")
                break
            except:
                continue
        
        if send_button:
            try:
                send_button.click()
                log_message("Botão de enviar clicado")
            except Exception as e:
                log_message(f"Erro ao clicar no botão de enviar: {e}")
                try:
                    actions = webdriver.ActionChains(driver)
                    actions.move_to_element(send_button).click().perform()
                    log_message("Botão de enviar clicado via ActionChains")
                except:
                    log_message("Tentando enviar com ENTER...")
                    message_box.send_keys(Keys.ENTER)
        else:
            # Se não encontrar o botão, tentar com ENTER
            log_message("Botão de enviar não encontrado, tentando com ENTER...")
            message_box.send_keys(Keys.ENTER)
            log_message("Tecla ENTER enviada")
        
        # Aguardar para confirmar envio
        time.sleep(5)
        
        # Verificar se a mensagem foi enviada
        try:
            # Procurar marcadores de mensagem enviada ou a própria mensagem no histórico
            sent_indicators = driver.find_elements(By.XPATH, "//span[@data-icon='msg-check']") + \
                              driver.find_elements(By.XPATH, "//span[@data-icon='msg-dcheck']") + \
                              driver.find_elements(By.XPATH, f"//div[contains(@class, 'message-out')]//*[contains(text(), '{mensagem[:20]}')]")
            
            if sent_indicators:
                log_message("Mensagem enviada com sucesso (confirmado)")
            else:
                log_message("Não foi possível confirmar o envio da mensagem, mas o processo foi concluído")
        except Exception as e:
            log_message(f"Erro ao verificar status da mensagem: {e}")
        
        log_message("Processo de envio de mensagem concluído")
        return True
        
    except Exception as e:
        log_message(f"Problema ao enviar mensagem no WhatsApp: {e}")
        if driver:
            try:
                driver.save_screenshot("whatsapp_error_final.png")
            except:
                pass
        return False
        
    finally:
        if driver:
            driver.quit()

def run(token, accounts, group, logs, date_range='today', start_date=None, end_date=None, low_profit=None, high_profit=None, realloc_pct=None, abo_accounts=None):
    """
    Função principal com suporte a ABO
    """
    global ACCESS_TOKEN, AD_ACCOUNTS, WHATSAPP_GROUP, DATE_PRESET, LIMITE_LUCRO_BAIXO, LIMITE_LUCRO_ALTO, PERCENTUAL_REALOCACAO, logs_list, ABO_ACCOUNTS
    
    ACCESS_TOKEN = token
    AD_ACCOUNTS = accounts if accounts is not None else []
    WHATSAPP_GROUP = group
    logs_list = logs
    
    # Definir contas ABO se fornecidas
    if abo_accounts:
        ABO_ACCOUNTS = abo_accounts
    
    # Limpar dados de campanhas anteriores
    global campanhas_completas_data
    campanhas_completas_data = {}
    
    if date_range == 'custom' and start_date and end_date:
        DATE_PRESET = None
    else:
        presets = {'today': 'today', 'yesterday': 'yesterday', 'last7': 'last_7d'}
        DATE_PRESET = presets.get(date_range, 'today')
    
    if low_profit is not None:
        LIMITE_LUCRO_BAIXO = float(low_profit)
    if high_profit is not None:
        LIMITE_LUCRO_ALTO = float(high_profit)
    if realloc_pct is not None:
        PERCENTUAL_REALOCACAO = float(realloc_pct) if float(realloc_pct) <= 1.0 else float(realloc_pct)/100.0
    
    log_message("Iniciando realocação com: Token=" + token[:5] + "..." + token[-5:] + 
               f", Contas={accounts}, Grupo={group}")
    log_message(f"Parâmetros: Date Range={date_range}, Low Profit={low_profit}, High Profit={high_profit}, Realloc %={realloc_pct}")
    log_message(f"Contas ABO configuradas: {ABO_ACCOUNTS}")
    
    limpar_planilha()
    todas_campanhas = []
    
    for ad_account in AD_ACCOUNTS:
        log_message(f"Processando conta de anúncio: {ad_account}")
        log_message(f"Buscando campanhas para conta {ad_account}...")
        
        campaigns_url = f"https://graph.facebook.com/v17.0/{ad_account}/campaigns?fields=id,name,daily_budget,status&access_token={ACCESS_TOKEN}"
        
        if date_range == 'custom' and start_date and end_date:
            insights_url = (
                f"https://graph.facebook.com/v17.0/{ad_account}/insights?fields=campaign_id,campaign_name,spend,action_values"
                f"&time_range[since]={start_date}&time_range[until]={end_date}&level=campaign&access_token={ACCESS_TOKEN}"
            )
        else:
            insights_url = (
                f"https://graph.facebook.com/v17.0/{ad_account}/insights?fields=campaign_id,campaign_name,spend,action_values"
                f"&date_preset={DATE_PRESET}&level=campaign&access_token={ACCESS_TOKEN}"
            )
        
        campaigns = buscar_todos_dados_facebook(campaigns_url)
        log_message(f"Encontradas {len(campaigns)} campanhas.")
        
        insights = buscar_todos_dados_facebook(insights_url)
        log_message(f"Encontrados {len(insights)} insights.")
        
        # Processar campanhas com suporte a ABO
        campanhas_processadas = processar_dados_campanhas(
            campaigns, insights, ad_account,
            DATE_PRESET if DATE_PRESET else None,
            start_date if date_range == 'custom' else None,
            end_date if date_range == 'custom' else None
        )
        
        log_message(f"Processadas {len(campanhas_processadas)} campanhas ativas.")
        
        todas_campanhas.extend(campanhas_processadas)
    
    log_message(f"Total de {len(todas_campanhas)} campanhas ativas encontradas.")
    
    # Contar campanhas por tipo
    campanhas_cbo = [c for c in todas_campanhas if c.get("tipo_campanha") == "CBO"]
    campanhas_abo = [c for c in todas_campanhas if c.get("tipo_campanha") == "ABO"]
    log_message(f"Campanhas CBO: {len(campanhas_cbo)}, Campanhas ABO: {len(campanhas_abo)}")
    
    salvar_campanhas_excel(todas_campanhas)
    
    try:
        resultado = realocar_orcamentos()
        return resultado
    except Exception as e:
        log_message(f"Erro durante o processo de realocação: {e}")
        return False

if __name__ == "__main__":
    limpar_planilha()
    todas_campanhas = []
    
    for ad_account in AD_ACCOUNTS:
        log_message(f"[INFO] Processando conta de anúncio: {ad_account}")
        campaigns_url = f"https://graph.facebook.com/v17.0/{ad_account}/campaigns?fields=id,name,daily_budget,status&access_token={ACCESS_TOKEN}"
        insights_url = f"https://graph.facebook.com/v17.0/{ad_account}/insights?fields=campaign_id,campaign_name,spend,action_values&date_preset={DATE_PRESET}&level=campaign&access_token={ACCESS_TOKEN}"
        
        campaigns = buscar_todos_dados_facebook(campaigns_url)
        insights = buscar_todos_dados_facebook(insights_url)
        
        campanhas_processadas = processar_dados_campanhas(
            campaigns, insights, ad_account, DATE_PRESET
        )
        todas_campanhas.extend(campanhas_processadas)
    
    salvar_campanhas_excel(todas_campanhas)
    realocar_orcamentos()