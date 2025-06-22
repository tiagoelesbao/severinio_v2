import json
import os
import requests
import openpyxl
import time
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.FileHandler("escala_lucro.log"),
        logging.StreamHandler()
    ]
)

logs_list = None

def log_message(msg):
    mensagem = f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(mensagem)
    with open("run_log.txt", "a", encoding="utf-8") as f:
        f.write(mensagem + "\n")
    if logs_list is not None:
        logs_list.append(mensagem)
    logging.info(msg)

# Leitura de config.json
CONFIG_FILE = "config.json"
if not os.path.exists(CONFIG_FILE):
    default_config = {
        "ACCESS_TOKEN": "",
        "AD_ACCOUNTS": [],
        "ABO_ACCOUNTS": [],  # Nova configuração para contas ABO
        "LIMITE_LUCRO": 1,
        "VALOR_TOTAL_ESCALA": 1000,
        "MINIMO_ORCAMENTO": 100,
        "MAXIMO_ORCAMENTO": 10000,  # Nova configuração
        "WHATSAPP_GROUP": "#ZIP - ROAS IMPERIO",
        "DATE_PRESET": "today"
    }
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(default_config, f, indent=4, ensure_ascii=False)

with open(CONFIG_FILE, "r", encoding="utf-8") as f:
    config = json.load(f)

ACCESS_TOKEN = config.get("ACCESS_TOKEN", "")
AD_ACCOUNTS = config.get("AD_ACCOUNTS", [])
ABO_ACCOUNTS = config.get("ABO_ACCOUNTS", [])  # Contas que usam ABO
SPREADSHEET_PATH = "campanhas_lucro.xlsx"
LIMITE_LUCRO = float(config.get("LIMITE_LUCRO", 1))
VALOR_TOTAL_ESCALA = float(config.get("VALOR_TOTAL_ESCALA", 10000))
MINIMO_ORCAMENTO = float(config.get("MINIMO_ORCAMENTO", 100))
MAXIMO_ORCAMENTO = float(config.get("MAXIMO_ORCAMENTO", 10000))
WHATSAPP_GROUP = config.get("WHATSAPP_GROUP", "#ZIP - ROAS IMPERIO")
DATE_PRESET = config.get("DATE_PRESET", "today")

# Armazena dados completos das campanhas para uso no escalonamento
campanhas_completas_data = {}

def criar_planilha():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "CAMPANHAS"
    sheet.append([
        "ID da Conta de Anúncio", "ID da Campanha", "Nome da Campanha", "Tipo",
        "Orçamento Diário", "Gasto", "Valor de Conversões",
        "ROAS", "Lucro", "Novo Orçamento", "Detalhes AdSets"
    ])
    workbook.save(SPREADSHEET_PATH)
    log_message("Planilha criada com sucesso.")

def limpar_planilha():
    if os.path.exists(SPREADSHEET_PATH):
        try:
            workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
            if "CAMPANHAS" in workbook.sheetnames:
                sheet = workbook["CAMPANHAS"]
                sheet.delete_rows(2, sheet.max_row)
            else:
                workbook.create_sheet("CAMPANHAS")
            workbook.save(SPREADSHEET_PATH)
            log_message("Planilha limpa no início da execução.")
        except Exception as e:
            log_message(f"[ERRO] Falha ao limpar a planilha: {e}")
    else:
        log_message("Planilha não encontrada. Será criada uma nova ao salvar os dados.")
        criar_planilha()

def buscar_dados_facebook(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        log_message(f"[ERRO] Falha ao buscar dados do Facebook: {e}\nResposta: {response.text if 'response' in locals() else ''}")
        return {}

def buscar_todos_dados_facebook(url):
    todos_dados = []
    while url:
        dados = buscar_dados_facebook(url)
        if "error" in dados:
            log_message(f"[ERRO] Graph API retornou: {dados['error']}")
            break
        page_data = dados.get("data", [])
        todos_dados.extend(page_data)
        url = dados.get("paging", {}).get("next")
    return todos_dados

def detectar_tipo_campanha(campanha, ad_account):
    """
    Detecta se a campanha é CBO ou ABO
    CBO tem daily_budget > 0, ABO tem daily_budget = 0 ou null
    """
    daily_budget = campanha.get("daily_budget", 0)
    
    # Verificação adicional pela conta
    if ad_account in ABO_ACCOUNTS:
        return "ABO"
    
    # Verificação pelo orçamento
    if daily_budget and int(daily_budget) > 0:
        return "CBO"
    else:
        return "ABO"

def buscar_adsets_campanha(campaign_id):
    """Busca todos os ad sets de uma campanha ABO"""
    url = (
        f"https://graph.facebook.com/v17.0/{campaign_id}/adsets"
        f"?fields=id,name,daily_budget,status&access_token={ACCESS_TOKEN}"
    )
    return buscar_todos_dados_facebook(url)

def buscar_insights_adset(ad_account, campaign_id, date_preset=None, start_date=None, end_date=None):
    """Busca insights no nível de ad set para campanhas ABO"""
    # CORREÇÃO: Usar campaign.id ao invés de campaign_id
    filtering = f'[{{"field":"campaign.id","operator":"EQUAL","value":"{campaign_id}"}}]'
    
    if date_preset:
        url = (
            f"https://graph.facebook.com/v17.0/{ad_account}/insights"
            f"?fields=adset_id,adset_name,campaign_id,spend,action_values"
            f"&date_preset={date_preset}&level=adset"
            f"&filtering={filtering}"
            f"&access_token={ACCESS_TOKEN}"
        )
    else:
        url = (
            f"https://graph.facebook.com/v17.0/{ad_account}/insights"
            f"?fields=adset_id,adset_name,campaign_id,spend,action_values"
            f"&time_range[since]={start_date}&time_range[until]={end_date}"
            f"&level=adset"
            f"&filtering={filtering}"
            f"&access_token={ACCESS_TOKEN}"
        )
    
    return buscar_todos_dados_facebook(url)

def processar_campanha_abo(campanha, ad_account, date_preset=None, start_date=None, end_date=None):
    """
    Processa campanhas ABO agregando dados de todos os ad sets ativos
    """
    campaign_id = campanha["id"]
    log_message(f"Processando campanha ABO: {campanha['name']}")
    
    adsets = buscar_adsets_campanha(campaign_id)
    insights_adsets = buscar_insights_adset(ad_account, campaign_id, date_preset, start_date, end_date)
    
    # Agregar dados de todos os ad sets ativos
    total_orcamento = 0
    total_gasto = 0
    total_conversao = 0
    adsets_info = []
    adsets_ativos = 0
    
    for adset in adsets:
        if adset.get("status") == "ACTIVE":
            adsets_ativos += 1
            adset_id = adset["id"]
            daily_budget = float(adset.get("daily_budget", 0)) / 100
            total_orcamento += daily_budget
            
            # Buscar insight correspondente
            insight = next((i for i in insights_adsets if i.get("adset_id") == adset_id), None)
            
            if insight:
                gasto = float(insight.get("spend", 0))
                valor_conversao = sum(
                    float(a.get("value", 0))
                    for a in insight.get("action_values", [])
                    if a.get("action_type") in ['offsite_conversion.purchase', 'offsite_conversion.fb_pixel_purchase']
                )
                total_gasto += gasto
                total_conversao += valor_conversao
            else:
                gasto = 0
                valor_conversao = 0
            
            adsets_info.append({
                "adset_id": adset_id,
                "adset_name": adset["name"],
                "daily_budget": daily_budget,
                "gasto": gasto,
                "valor_conversao": valor_conversao
            })
    
    log_message(f"Campanha ABO {campaign_id}: {adsets_ativos} adsets ativos, orçamento total: R$ {total_orcamento:.2f}")
    
    lucro = total_conversao - total_gasto
    roas = round(total_conversao / total_gasto, 2) if total_gasto > 0 else 0
    
    return {
        "id_conta": ad_account,
        "id_campanha": campaign_id,
        "nome_campanha": campanha["name"],
        "tipo_campanha": "ABO",
        "orcamento_diario": total_orcamento,
        "gasto": total_gasto,
        "valor_conversao": total_conversao,
        "roas": roas,
        "lucro": lucro,
        "adsets_info": adsets_info,
        "detalhes_adsets": f"{adsets_ativos} adsets ativos"
    }

def processar_dados_campanhas(campanhas, insights, ad_account, date_preset=None, start_date=None, end_date=None):
    """
    Processa dados das campanhas, detectando automaticamente se são CBO ou ABO
    """
    campanhas_filtradas = []
    
    for campanha in campanhas:
        if campanha.get("status", "").upper().strip() == "ACTIVE":
            tipo_campanha = detectar_tipo_campanha(campanha, ad_account)
            
            if tipo_campanha == "ABO":
                # Processar como ABO
                dados_campanha = processar_campanha_abo(campanha, ad_account, date_preset, start_date, end_date)
                campanhas_filtradas.append(dados_campanha)
                # Armazenar dados completos para uso posterior
                campanhas_completas_data[campanha["id"]] = dados_campanha
            else:
                # Processar como CBO (código original)
                insight = next((i for i in insights if i.get("campaign_id") == campanha.get("id")), None)
                
                if insight:
                    gasto = float(insight.get("spend", 0))
                    valor_conversao = sum(
                        float(a.get("value", 0))
                        for a in insight.get("action_values", [])
                        if a.get("action_type") in ['offsite_conversion.purchase', 'offsite_conversion.fb_pixel_purchase']
                    )
                else:
                    gasto = 0.0
                    valor_conversao = 0.0
                
                daily_budget = float(campanha.get("daily_budget", 0)) / 100
                lucro = valor_conversao - gasto
                roas = round(valor_conversao / gasto, 2) if gasto > 0 else 0
                
                dados_campanha = {
                    "id_conta": ad_account,
                    "id_campanha": campanha["id"],
                    "nome_campanha": campanha["name"],
                    "tipo_campanha": "CBO",
                    "orcamento_diario": daily_budget,
                    "gasto": gasto,
                    "valor_conversao": valor_conversao,
                    "roas": roas,
                    "lucro": lucro,
                    "adsets_info": None,
                    "detalhes_adsets": "N/A"
                }
                
                campanhas_filtradas.append(dados_campanha)
                campanhas_completas_data[campanha["id"]] = dados_campanha
    
    return campanhas_filtradas

def salvar_campanhas_excel(campanhas):
    if not os.path.exists(SPREADSHEET_PATH):
        criar_planilha()
    
    try:
        workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
        sheet = workbook["CAMPANHAS"]
        
        for campanha in campanhas:
            sheet.append([
                campanha["id_conta"],
                campanha["id_campanha"],
                campanha["nome_campanha"],
                campanha.get("tipo_campanha", "CBO"),
                campanha["orcamento_diario"],
                campanha["gasto"],
                campanha["valor_conversao"],
                campanha["roas"],
                campanha["lucro"],
                "",  # Novo orçamento
                campanha.get("detalhes_adsets", "")
            ])
        
        workbook.save(SPREADSHEET_PATH)
        log_message("Dados das campanhas salvos na planilha.")
    except Exception as e:
        log_message(f"[ERRO] Falha ao salvar planilha: {e}")

def calcular_orcamento_total():
    try:
        workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
        sheet = workbook["CAMPANHAS"]
        total = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            novo_orcamento = row[9] if row[9] is not None else row[4]
            total += novo_orcamento
        return total
    except Exception as e:
        log_message(f"[ERRO] Falha ao calcular orçamento total: {e}")
        return 0

def atualizar_orcamento_adset(adset_id, novo_orcamento):
    """Atualiza o orçamento de um ad set específico"""
    url = f"https://graph.facebook.com/v17.0/{adset_id}"
    payload = {
        "daily_budget": int(novo_orcamento * 100),
        "access_token": ACCESS_TOKEN
    }
    
    try:
        log_message(f"Atualizando orçamento do AdSet {adset_id} para R$ {novo_orcamento:.2f}")
        response = requests.post(url, data=payload)
        result = response.json()
        
        if result.get("success"):
            log_message(f"Orçamento do AdSet {adset_id} atualizado com sucesso")
            return True
        else:
            erro_msg = result.get('error', {}).get('message', 'Erro desconhecido')
            log_message(f"[ERRO] Falha ao atualizar AdSet {adset_id}: {erro_msg}")
            return False
    except requests.exceptions.RequestException as e:
        log_message(f"[ERRO] Erro na requisição para atualizar AdSet {adset_id}: {e}")
        return False

def atualizar_orcamento_facebook(id_campanha, novo_orcamento):
    """Atualiza o orçamento de campanhas CBO"""
    url = f"https://graph.facebook.com/v17.0/{id_campanha}"
    payload = {
        "daily_budget": int(novo_orcamento * 100),
        "access_token": ACCESS_TOKEN
    }
    try:
        log_message(f"Enviando atualização de orçamento para campanha {id_campanha}")
        response = requests.post(url, data=payload)
        result = response.json()
        log_message(f"Resposta da API: {result}")
        if result.get("success"):
            log_message(f"Orçamento atualizado para a campanha {id_campanha}: R$ {novo_orcamento:.2f}")
            return True
        else:
            log_message(f"[ERRO] Falha ao atualizar campanha {id_campanha}: {result.get('error', {}).get('message')}")
            return False
    except requests.exceptions.RequestException as e:
        log_message(f"[ERRO] Erro na requisição para atualizar orçamento: {e}")
        return False

def escalar_campanhas():
    if not os.path.exists(SPREADSHEET_PATH):
        log_message("[ERRO] Planilha de campanhas não encontrada.")
        return False
    
    workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
    if "CAMPANHAS" not in workbook.sheetnames:
        log_message("[ERRO] Aba 'CAMPANHAS' não encontrada na planilha.")
        return False
    
    sheet = workbook["CAMPANHAS"]
    
    # Lista unificada de unidades escaláveis (campanhas CBO + AdSets ABO individuais)
    unidades_escalaveis = []
    
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        lucro = row[8]
        tipo_campanha = row[3] if len(row) > 3 else "CBO"
        id_campanha = row[1]
        
        if tipo_campanha == "ABO":
            # IMPORTANTE: Para campanhas ABO, SEMPRE processar os AdSets
            # independentemente do lucro da campanha
            campanha_completa = campanhas_completas_data.get(id_campanha)
            if campanha_completa and campanha_completa.get("adsets_info"):
                log_message(f"Processando AdSets da campanha ABO: {row[2]} (lucro campanha: R$ {lucro:.2f})")
                adsets_adicionados = 0
                
                for adset in campanha_completa["adsets_info"]:
                    adset_lucro = adset.get('valor_conversao', 0) - adset.get('gasto', 0)
                    # Aplicar filtro de lucro apenas no nível do AdSet
                    if adset_lucro >= LIMITE_LUCRO:
                        unidades_escalaveis.append({
                            "tipo": "ABO_ADSET",
                            "linha_index": row_index,
                            "id_campanha": id_campanha,
                            "id_adset": adset['adset_id'],
                            "nome": f"{row[2]} - {adset['adset_name']}",
                            "nome_campanha": row[2],
                            "orcamento_atual": adset['daily_budget'],
                            "lucro": adset_lucro,
                            "adset_info": adset,
                            "campanha_info": campanha_completa
                        })
                        adsets_adicionados += 1
                
                log_message(f"  -> {adsets_adicionados} AdSets com lucro >= R$ {LIMITE_LUCRO:.2f}")
        else:
            # Para campanhas CBO, aplicar filtro de lucro na campanha
            if lucro is not None and lucro >= LIMITE_LUCRO:
                unidades_escalaveis.append({
                    "tipo": "CBO",
                    "linha_index": row_index,
                    "id_campanha": id_campanha,
                    "nome": row[2],
                    "nome_campanha": row[2],
                    "orcamento_atual": row[4],
                    "lucro": lucro
                })
    
    if not unidades_escalaveis:
        log_message("[INFO] Nenhuma unidade para escalar.")
        return False
    
    # Ordenar por lucro (maior primeiro) para priorizar os melhores
    unidades_escalaveis.sort(key=lambda x: x["lucro"], reverse=True)
    
    # Calcular soma total dos lucros
    soma_lucro = sum(u["lucro"] for u in unidades_escalaveis)
    
    log_message(f"[INFO] {len(unidades_escalaveis)} unidades para escalonamento:")
    log_message(f"- Campanhas CBO: {sum(1 for u in unidades_escalaveis if u['tipo'] == 'CBO')}")
    log_message(f"- AdSets ABO: {sum(1 for u in unidades_escalaveis if u['tipo'] == 'ABO_ADSET')}")
    log_message(f"- Soma dos lucros: R$ {soma_lucro:.2f}")
    
    # Distribuir verba proporcionalmente
    unidades_escaladas = []
    total_distribuido = 0
    campanhas_modificadas = {}  # Para rastrear mudanças nas campanhas ABO
    
    for unidade in unidades_escalaveis:
        # Calcular incremento proporcional ao lucro
        if soma_lucro > 0:
            proporcao = unidade["lucro"] / soma_lucro
            incremento = VALOR_TOTAL_ESCALA * proporcao
        else:
            incremento = VALOR_TOTAL_ESCALA / len(unidades_escalaveis)
        
        # Aplicar limites mínimos de incremento
        if incremento < 10:  # Incremento mínimo de R$ 10
            continue
            
        sucesso = False
        
        if unidade["tipo"] == "CBO":
            # Escalar campanha CBO
            novo_orcamento = min(
                max(unidade["orcamento_atual"] + incremento, MINIMO_ORCAMENTO),
                MAXIMO_ORCAMENTO
            )
            
            # Calcular incremento real (pode ser menor devido aos limites)
            incremento_real = novo_orcamento - unidade["orcamento_atual"]
            
            if atualizar_orcamento_facebook(unidade["id_campanha"], novo_orcamento):
                sheet.cell(row=unidade["linha_index"], column=10).value = novo_orcamento
                unidades_escaladas.append(f"{unidade['nome']} (CBO) +R$ {incremento_real:.2f}")
                total_distribuido += incremento_real
                sucesso = True
                log_message(f"Campanha CBO {unidade['id_campanha']} escalada para R$ {novo_orcamento:.2f} (+R$ {incremento_real:.2f})")
                
        else:  # ABO_ADSET
            # Escalar AdSet individual
            adset = unidade["adset_info"]
            novo_orcamento = min(
                max(adset['daily_budget'] + incremento, MINIMO_ORCAMENTO),
                MAXIMO_ORCAMENTO
            )
            
            # Calcular incremento real
            incremento_real = novo_orcamento - adset['daily_budget']
            
            if atualizar_orcamento_adset(unidade["id_adset"], novo_orcamento):
                # Rastrear mudança total na campanha
                if unidade["id_campanha"] not in campanhas_modificadas:
                    campanhas_modificadas[unidade["id_campanha"]] = {
                        "linha_index": unidade["linha_index"],
                        "orcamento_original": unidade["campanha_info"]["orcamento_diario"],
                        "incremento_total": 0,
                        "nome": unidade["nome_campanha"]
                    }
                
                campanhas_modificadas[unidade["id_campanha"]]["incremento_total"] += incremento_real
                
                unidades_escaladas.append(f"{unidade['nome']} (ABO AdSet) +R$ {incremento_real:.2f}")
                total_distribuido += incremento_real
                sucesso = True
                log_message(f"AdSet {unidade['id_adset']} escalado de R$ {adset['daily_budget']:.2f} para R$ {novo_orcamento:.2f} (+R$ {incremento_real:.2f})")
    
    # Atualizar orçamentos totais das campanhas ABO na planilha
    for id_campanha, info in campanhas_modificadas.items():
        novo_orcamento_total = info["orcamento_original"] + info["incremento_total"]
        sheet.cell(row=info["linha_index"], column=10).value = novo_orcamento_total
        log_message(f"Campanha ABO {info['nome']} - orçamento total atualizado para R$ {novo_orcamento_total:.2f}")
    
    workbook.save(SPREADSHEET_PATH)
    total_orcamento_atual = calcular_orcamento_total()
    
    # Criar mensagem detalhada
    mensagem = (
        f"✅ Escala realizada com sucesso!\n\n"
        f"💰 Total distribuído: R$ {total_distribuido:.2f}\n"
        f"📊 Unidades escaladas: {len(unidades_escaladas)}\n"
        f"📈 Orçamento total atual: R$ {total_orcamento_atual:.2f}\n\n"
    )
    
    # Separar por tipo
    campanhas_cbo_escaladas = [u for u in unidades_escaladas if "(CBO)" in u]
    adsets_abo_escalados = [u for u in unidades_escaladas if "(ABO AdSet)" in u]
    
    if campanhas_cbo_escaladas:
        mensagem += f"🎯 Campanhas CBO ({len(campanhas_cbo_escaladas)}):\n"
        for i, campanha in enumerate(campanhas_cbo_escaladas[:5]):  # Top 5 CBO
            mensagem += f"{i+1}. {campanha}\n"
        if len(campanhas_cbo_escaladas) > 5:
            mensagem += f"... e mais {len(campanhas_cbo_escaladas) - 5} campanhas CBO\n"
        mensagem += "\n"
    
    if adsets_abo_escalados:
        mensagem += f"🔄 AdSets ABO ({len(adsets_abo_escalados)}):\n"
        for i, adset in enumerate(adsets_abo_escalados[:5]):  # Top 5 ABO
            mensagem += f"{i+1}. {adset}\n"
        if len(adsets_abo_escalados) > 5:
            mensagem += f"... e mais {len(adsets_abo_escalados) - 5} AdSets ABO\n"
    
    # Log detalhado
    log_message(f"[RESUMO] Total de unidades escaladas: {len(unidades_escaladas)}")
    log_message(f"[RESUMO] Campanhas CBO escaladas: {len(campanhas_cbo_escaladas)}")
    log_message(f"[RESUMO] AdSets ABO escalados: {len(adsets_abo_escalados)}")
    log_message(f"[RESUMO] Total distribuído: R$ {total_distribuido:.2f}")
    
    sucesso_whatsapp = enviar_mensagem_whatsapp(WHATSAPP_GROUP, mensagem)
    
    if not sucesso_whatsapp:
        log_message(f"[AVISO] Falha ao enviar mensagem WhatsApp, mas a escala foi concluída: {mensagem}")
    else:
        log_message(f"[INFO] Mensagem enviada com sucesso")
    
    log_message("Processo de escala concluído com sucesso!")
    return True

def limpar_mensagem_whatsapp(mensagem):
    """
    Remove caracteres especiais e emojis que podem causar problemas no WhatsApp Web
    """
    import re
    
    # Substituir emojis específicos por texto
    substituicoes = {
        '✅': '[OK]',
        '💰': '[$$]',
        '📊': '[DADOS]',
        '📈': '[SUBIU]',
        '📉': '[DESCEU]',
        '🔥': '[HOT]',
        '🟡': '[*]',
        '🌎': '[MUNDO]',
        '•': '-'
    }
    
    for emoji, texto in substituicoes.items():
        mensagem = mensagem.replace(emoji, texto)
    
    # Remover outros caracteres Unicode problemáticos
    # Mantém apenas caracteres ASCII e alguns latinos comuns
    mensagem = ''.join(char for char in mensagem if ord(char) < 256)
    
    return mensagem

def enviar_mensagem_whatsapp(grupo, mensagem_original):
    """
    Função para enviar mensagens para um grupo do WhatsApp
    """
    # Limpar mensagem antes de enviar
    mensagem = limpar_mensagem_whatsapp(mensagem_original)
    
    log_message(f"Iniciando envio de mensagem para o grupo: {grupo}")
    log_message(f"Mensagem limpa: {mensagem[:100]}...")  # Log dos primeiros 100 caracteres
    
    driver = None
    
    try:
        # Configuração do navegador Chrome/Brave
        brave_options = Options()
        brave_options.binary_location = r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"
        brave_options.add_argument(r"--user-data-dir=C:\Users\Pichau\AppData\Local\BraveSoftware\Brave-Browser\User Data")
        brave_options.add_argument(r"--profile-directory=Default")
        brave_options.add_argument("--start-maximized")
        brave_options.add_argument("--window-size=1920,1080")
        
        # Adicionar opção para evitar detecção de automação
        brave_options.add_argument("--disable-blink-features=AutomationControlled")
        brave_options.add_argument("--disable-extensions")
        brave_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        brave_options.add_experimental_option("useAutomationExtension", False)
        
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=brave_options)
        
        # Remover flag navigator.webdriver
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            })
            """
        })
        
        # Abrir WhatsApp Web
        log_message("Abrindo WhatsApp Web...")
        driver.get("https://web.whatsapp.com")
        
        # Múltiplos seletores para verificar se o WhatsApp Web carregou
        whatsapp_loaded_selectors = [
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Caixa de texto de pesquisa']"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-placeholder='Pesquisar ou começar uma nova conversa']"),
            (By.XPATH, "//div[@role='textbox' and contains(@aria-label, 'Pesquisar')]"),
            (By.XPATH, "//div[contains(@class, '_ai04')]"),
            (By.XPATH, "//div[@id='pane-side']")
        ]
        
        # Aguardar carregamento da página com múltiplos seletores
        log_message("Aguardando carregamento do WhatsApp Web...")
        carregou = False
        for selector_type, selector_value in whatsapp_loaded_selectors:
            try:
                WebDriverWait(driver, 60).until(
                    EC.presence_of_element_located((selector_type, selector_value))
                )
                log_message(f"WhatsApp Web carregado, elemento detectado: {selector_value}")
                carregou = True
                break
            except:
                continue
        
        if not carregou:
            log_message("Verificando se o WhatsApp carregou usando outro método...")
            try:
                WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@data-testid='chat-list']"))
                )
                log_message("WhatsApp Web carregado (detectado via chat-list)")
                carregou = True
            except:
                pass
        
        if not carregou:
            log_message("Falha ao detectar carregamento do WhatsApp Web")
            driver.save_screenshot("whatsapp_nao_carregou.png")
            return False
        
        # Garantir que a página está completamente carregada
        time.sleep(2)
        
        # Múltiplos seletores para o campo de pesquisa
        search_selectors = [
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Caixa de texto de pesquisa']"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-placeholder='Pesquisar ou começar uma nova conversa']"),
            (By.XPATH, "//div[@role='textbox' and contains(@aria-label, 'Pesquisar')]"),
            (By.XPATH, "//button[@aria-label='Pesquisar ou começar uma nova conversa']"),
            (By.XPATH, "//div[contains(@class, 'x10l6tqk')]"),
            (By.XPATH, "//div[contains(@class, 'lexical-rich-text-input')]//div[@role='textbox']")
        ]
        
        # Encontrar o campo de pesquisa
        search_element = None
        for selector_type, selector_value in search_selectors:
            try:
                search_element = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((selector_type, selector_value))
                )
                log_message(f"Campo de pesquisa encontrado com seletor: {selector_value}")
                break
            except:
                continue
        
        if not search_element:
            log_message("Campo de pesquisa não encontrado")
            driver.save_screenshot("campo_pesquisa_nao_encontrado.png")
            return False
        
        # Clicar no campo de pesquisa e inserir o nome do grupo
        try:
            # Usando clique normal primeiro
            search_element.click()
            time.sleep(1)
            search_element.clear()
            search_element.send_keys(grupo)
            log_message(f"Texto de pesquisa '{grupo}' inserido.")
        except Exception as e:
            log_message(f"Erro ao clicar no campo de pesquisa normalmente: {e}")
            try:
                # Se falhar, tente com ActionChains
                actions = webdriver.ActionChains(driver)
                actions.move_to_element(search_element).click().perform()
                time.sleep(1)
                search_element.clear()
                search_element.send_keys(grupo)
                log_message(f"Texto de pesquisa '{grupo}' inserido via ActionChains.")
            except Exception as e2:
                log_message(f"Falha ao inserir texto de pesquisa: {e2}")
                driver.save_screenshot("erro_inserir_pesquisa.png")
                return False
        
        # Aguardar resultados da pesquisa
        time.sleep(3)
        
        # Múltiplos seletores para encontrar o grupo
        group_selectors = [
            (By.XPATH, f"//span[@title='{grupo}']"),
            (By.XPATH, f"//span[contains(text(), '{grupo}')]"),
            (By.XPATH, f"//div[contains(text(), '{grupo}')]")
        ]
        
        # Encontrar e clicar no grupo
        group_element = None
        for selector_type, selector_value in group_selectors:
            try:
                group_element = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((selector_type, selector_value))
                )
                log_message(f"Grupo encontrado com seletor: {selector_value}")
                break
            except:
                continue
        
        if not group_element:
            log_message("Grupo não encontrado")
            driver.save_screenshot("grupo_nao_encontrado.png")
            return False
        
        # Clicar no grupo usando clique normal
        try:
            group_element.click()
            log_message("Grupo clicado com sucesso")
        except Exception as e:
            log_message(f"Erro ao clicar no grupo normalmente: {e}")
            try:
                # Se falhar, tente com ActionChains
                actions = webdriver.ActionChains(driver)
                actions.move_to_element(group_element).click().perform()
                log_message("Grupo clicado com ActionChains")
            except Exception as e2:
                log_message(f"Falha ao clicar no grupo: {e2}")
                driver.save_screenshot("erro_clique_grupo.png")
                return False
        
        # Verificar se o grupo foi realmente clicado
        time.sleep(3)
        try:
            # Verificar se estamos na conversa, procurando por elemento que só aparece depois de clicar no grupo
            chat_header = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, f"//span[contains(text(), '{grupo}')]/ancestor::div[contains(@role, 'button')]"))
            )
            log_message("Conversa do grupo aberta com sucesso")
        except:
            log_message("Não foi possível confirmar se a conversa do grupo foi aberta")
            driver.save_screenshot("conversa_nao_confirmada.png")
            return False
        
        # Múltiplos seletores para o campo de mensagem
        message_selectors = [
            (By.XPATH, "//div[@role='textbox' and @data-tab='10']"),
            (By.XPATH, "//div[@role='textbox' and contains(@aria-label, 'Digite uma mensagem')]"),
            (By.CSS_SELECTOR, "div[role='textbox'][data-tab='10']"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Digite uma mensagem']"),
            (By.XPATH, "//div[contains(@class, 'lexical-rich-text-input')]//div[@role='textbox']"),
            (By.XPATH, "//footer//div[@role='textbox']"),
            (By.XPATH, "//div[@data-testid='conversation-compose-box-input']"),
            (By.XPATH, "//div[@title='Digite uma mensagem']")
        ]
        
        # Encontrar o campo de mensagem
        log_message("Localizando campo de mensagem...")
        message_box = None
        for selector_type, selector_value in message_selectors:
            try:
                message_box = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((selector_type, selector_value))
                )
                log_message(f"Campo de mensagem encontrado com seletor: {selector_value}")
                break
            except:
                continue
        
        if not message_box:
            log_message("Campo de mensagem não encontrado")
            driver.save_screenshot("campo_mensagem_nao_encontrado.png")
            return False
        
        # Clicar e digitar a mensagem
        log_message("Tentando digitar mensagem...")
        try:
            message_box.click()
            time.sleep(1)
            message_box.clear()
            message_box.send_keys(mensagem)
            log_message("Mensagem digitada com sucesso")
        except Exception as e:
            log_message(f"Erro ao digitar mensagem normalmente: {e}")
            try:
                actions = webdriver.ActionChains(driver)
                actions.move_to_element(message_box).click().perform()
                time.sleep(1)
                message_box.send_keys(mensagem)
                log_message("Mensagem digitada via ActionChains")
            except Exception as e2:
                log_message(f"Falha ao digitar mensagem: {e2}")
                driver.save_screenshot("erro_digitacao.png")
                return False
        
        time.sleep(1)
        
        # Enviar a mensagem
        log_message("Enviando mensagem...")
        
        # Primeiro tenta encontrar o botão de enviar
        send_selectors = [
            (By.CSS_SELECTOR, "button[aria-label='Enviar']"),
            (By.CSS_SELECTOR, "button[data-tab='11']"),
            (By.XPATH, "//button[contains(@class, '_3wFFT')]"),
            (By.XPATH, "//button[@data-icon='send']"),
            (By.XPATH, "//span[@data-icon='send']"),
            (By.XPATH, "//button[@data-testid='send']"),
            (By.XPATH, "//button[@aria-label='Enviar mensagem']")
        ]
        
        send_button = None
        for selector_type, selector_value in send_selectors:
            try:
                send_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((selector_type, selector_value))
                )
                log_message(f"Botão de enviar encontrado com seletor: {selector_value}")
                break
            except:
                continue
        
        if send_button:
            try:
                send_button.click()
                log_message("Botão de enviar clicado")
            except Exception as e:
                log_message(f"Erro ao clicar no botão de enviar: {e}")
                try:
                    actions = webdriver.ActionChains(driver)
                    actions.move_to_element(send_button).click().perform()
                    log_message("Botão de enviar clicado via ActionChains")
                except:
                    log_message("Tentando enviar com ENTER...")
                    message_box.send_keys(Keys.ENTER)
        else:
            # Se não encontrar o botão, tentar com ENTER
            log_message("Botão de enviar não encontrado, tentando com ENTER...")
            message_box.send_keys(Keys.ENTER)
            log_message("Tecla ENTER enviada")
        
        # Aguardar para confirmar envio
        time.sleep(5)
        
        # Verificar se a mensagem foi enviada
        try:
            # Procurar marcadores de mensagem enviada ou a própria mensagem no histórico
            sent_indicators = driver.find_elements(By.XPATH, "//span[@data-icon='msg-check']") + \
                              driver.find_elements(By.XPATH, "//span[@data-icon='msg-dcheck']") + \
                              driver.find_elements(By.XPATH, f"//div[contains(@class, 'message-out')]//*[contains(text(), '{mensagem[:20]}')]")
            
            if sent_indicators:
                log_message("Mensagem enviada com sucesso (confirmado)")
            else:
                log_message("Não foi possível confirmar o envio da mensagem, mas o processo foi concluído")
        except Exception as e:
            log_message(f"Erro ao verificar status da mensagem: {e}")
        
        log_message("Processo de envio de mensagem concluído")
        return True
        
    except Exception as e:
        log_message(f"Problema ao enviar mensagem no WhatsApp: {e}")
        if driver:
            try:
                driver.save_screenshot("whatsapp_error_final.png")
            except:
                pass
        return False
        
    finally:
        if driver:
            driver.quit()

def run(token, accounts, group, logs, date_range='today', start_date=None, end_date=None, min_profit=None, scale_value=None, abo_accounts=None):
    """
    Função principal que executa o processo de escala de orçamento
    
    Args:
        token (str): Token de acesso da API do Facebook
        accounts (list): Lista de IDs de contas de anúncios
        group (str): Nome do grupo do WhatsApp para enviar notificações
        logs (list): Lista para armazenar logs
        date_range (str): Período de dados a considerar (today, yesterday, last7, custom)
        start_date (str): Data inicial (se date_range for custom)
        end_date (str): Data final (se date_range for custom)
        min_profit (float): Lucro mínimo para considerar escala
        scale_value (float): Valor total para escalar
        abo_accounts (list): Lista de contas que usam ABO
        
    Returns:
        bool: True se o processo foi concluído com sucesso, False caso contrário
    """
    global ACCESS_TOKEN, AD_ACCOUNTS, WHATSAPP_GROUP, DATE_PRESET, LIMITE_LUCRO, VALOR_TOTAL_ESCALA, logs_list, ABO_ACCOUNTS
    
    ACCESS_TOKEN = token
    AD_ACCOUNTS = accounts if accounts is not None else []
    WHATSAPP_GROUP = group
    logs_list = logs
    
    # Definir contas ABO se fornecidas
    if abo_accounts:
        ABO_ACCOUNTS = abo_accounts
        # IMPORTANTE: Adicionar contas ABO à lista de contas a processar
        # para garantir que sejam incluídas no processamento
        for conta_abo in abo_accounts:
            if conta_abo not in AD_ACCOUNTS:
                AD_ACCOUNTS.append(conta_abo)
                log_message(f"Conta ABO {conta_abo} adicionada à lista de processamento")
    
    # Limpar dados de campanhas anteriores
    global campanhas_completas_data
    campanhas_completas_data = {}
    
    if date_range == 'custom' and start_date and end_date:
        DATE_PRESET = None
    else:
        presets = {'today': 'today', 'yesterday': 'yesterday', 'last7': 'last_7d'}
        DATE_PRESET = presets.get(date_range, 'today')
    
    if min_profit is not None:
        LIMITE_LUCRO = float(min_profit)
    if scale_value is not None:
        VALOR_TOTAL_ESCALA = float(scale_value)
        
    log_message("Iniciando escala com: Token=" + token[:5] + "..." + token[-5:] + 
               f", Contas={AD_ACCOUNTS}, Grupo={group}")
    log_message(f"Parâmetros: Date Range={date_range}, Min Profit={min_profit}, Scale Value={scale_value}")
    log_message(f"Contas ABO configuradas: {ABO_ACCOUNTS}")
    
    try:
        limpar_planilha()
        todas_campanhas = []
        
        # Processar TODAS as contas (incluindo ABO)
        for ad_account in AD_ACCOUNTS:
            tipo_conta = "ABO" if ad_account in ABO_ACCOUNTS else "CBO"
            log_message(f"Processando conta de anúncio {tipo_conta}: {ad_account}")
            log_message(f"Buscando campanhas para conta {ad_account}...")
            
            campaigns_url = f"https://graph.facebook.com/v17.0/{ad_account}/campaigns?fields=id,name,daily_budget,status&access_token={ACCESS_TOKEN}"
            
            if date_range == 'custom' and start_date and end_date:
                insights_url = f"https://graph.facebook.com/v17.0/{ad_account}/insights?fields=campaign_id,campaign_name,spend,action_values&time_range[since]={start_date}&time_range[until]={end_date}&level=campaign&access_token={ACCESS_TOKEN}"
            else:
                insights_url = f"https://graph.facebook.com/v17.0/{ad_account}/insights?fields=campaign_id,campaign_name,spend,action_values&date_preset={DATE_PRESET}&level=campaign&access_token={ACCESS_TOKEN}"
            
            campaigns = buscar_todos_dados_facebook(campaigns_url)
            log_message(f"Encontradas {len(campaigns)} campanhas.")
            
            insights = buscar_todos_dados_facebook(insights_url)
            log_message(f"Encontrados {len(insights)} insights.")
            
            # Processar campanhas com suporte a ABO
            campanhas_processadas = processar_dados_campanhas(
                campaigns, insights, ad_account, 
                DATE_PRESET if DATE_PRESET else None, 
                start_date if date_range == 'custom' else None,
                end_date if date_range == 'custom' else None
            )
            
            log_message(f"Processadas {len(campanhas_processadas)} campanhas ativas.")
            
            # Log detalhado para campanhas ABO
            if ad_account in ABO_ACCOUNTS:
                campanhas_abo_desta_conta = [c for c in campanhas_processadas if c.get("tipo_campanha") == "ABO"]
                if campanhas_abo_desta_conta:
                    for camp in campanhas_abo_desta_conta:
                        log_message(f"  - Campanha ABO: {camp['nome_campanha']} com {len(camp.get('adsets_info', []))} adsets")
            
            todas_campanhas.extend(campanhas_processadas)
        
        log_message(f"Total de {len(todas_campanhas)} campanhas ativas encontradas.")
        
        # Contar campanhas por tipo
        campanhas_cbo = [c for c in todas_campanhas if c.get("tipo_campanha") == "CBO"]
        campanhas_abo = [c for c in todas_campanhas if c.get("tipo_campanha") == "ABO"]
        log_message(f"Campanhas CBO: {len(campanhas_cbo)}, Campanhas ABO: {len(campanhas_abo)}")
        
        # Log detalhado de campanhas ABO
        if campanhas_abo:
            total_adsets = sum(len(c.get("adsets_info", [])) for c in campanhas_abo)
            log_message(f"Total de AdSets em campanhas ABO: {total_adsets}")
        
        salvar_campanhas_excel(todas_campanhas)
        
        resultado = escalar_campanhas()
        return resultado
        
    except Exception as e:
        log_message(f"Erro durante o processo de escala: {e}")
        import traceback
        log_message(traceback.format_exc())
        return False

if __name__ == "__main__":
    limpar_planilha()
    todas_campanhas = []
    
    for ad_account in AD_ACCOUNTS:
        log_message(f"[INFO] Processando conta de anúncio: {ad_account}")
        campaigns_url = f"https://graph.facebook.com/v17.0/{ad_account}/campaigns?fields=id,name,daily_budget,status&access_token={ACCESS_TOKEN}"
        insights_url = f"https://graph.facebook.com/v17.0/{ad_account}/insights?fields=campaign_id,campaign_name,spend,action_values&date_preset={DATE_PRESET}&level=campaign&access_token={ACCESS_TOKEN}"
        
        campaigns = buscar_todos_dados_facebook(campaigns_url)
        insights = buscar_todos_dados_facebook(insights_url)
        
        campanhas_processadas = processar_dados_campanhas(
            campaigns, insights, ad_account, DATE_PRESET
        )
        todas_campanhas.extend(campanhas_processadas)
    
    salvar_campanhas_excel(todas_campanhas)
    escalar_campanhas()